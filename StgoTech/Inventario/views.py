from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
import openpyxl
from .forms import *
from django.db.models import Q , Sum
from django.http.response import JsonResponse
from django.contrib.auth import logout
from django.contrib.auth.views import LoginView
from django.http import HttpResponse
from django.db.models import Count
from django.db import connection
from Inventario.forms import OrdenConsumoForm
from django.contrib import messages
from datetime import datetime, timedelta
from django.db.models.functions import TruncMonth

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
# Vistas relacionadas al inicio y cierre de sesión
def redirect_login(request):
    return redirect('login')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'

def cerrar_sesion(request):
    logout(request)
    return redirect('/login')

def error_404(request, exception):
    return render(request, 'errores/404.html', status=404)

def error_400(request, exception):
    return render(request, 'errores/400.html', status=400)

def error_403(request, exception):
    return render(request, 'errores/403.html', status=403)

def error_500(request):
    return render(request, 'errores/500.html', status=500)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

#REDIRIGE A LA PAGINA PRINCIPAL INICIO
def index(request):
    return redirect('dashboard')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

#VISTA DASHBOARD
def dashboard(request):
    return render(request, 'formularios/dashboard.html')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def get_chart_data(request):
    priority_order = ['Alta', 'Media', 'Baja']
    data = Comat.objects.values('prioridad').annotate(count=Count('prioridad'))

    data = sorted(data, key=lambda item: priority_order.index(item['prioridad']))
    # Asigna colores según la prioridad
    # color_mapping = {
    #     'Alta': 'rgba(255, 0, 0, 0.5)',    # Rojo
    #     'Media': 'rgba(255, 255, 0, 0.5)', # Amarillo
    #     'Baja': 'rgba(0, 128, 0, 0.5)'     # Verde
    # }
    color_mapping = {
        'Alta': 'rgba(20, 66, 102, 0.8)',    # Rojo
        'Media': 'rgba(32, 144, 215, 0.8)', # Amarillo
        'Baja': 'rgba(192, 224, 247, 0.8)'     # Verde
    }
    for item in data:
        item['color'] = color_mapping.get(item['prioridad'], 'rgba(0, 0, 0, 0.6)')  # Por defecto, negro

    return JsonResponse(list(data), safe=False)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def get_pks_by_priority(request, priority):
    stdf_pk = Comat.objects.filter(prioridad=priority).values_list('stdf_pk', flat=True)
    return JsonResponse({'stdf_pk': list(stdf_pk)}, safe=False)
# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def get_chart_data_repuesto_owner(request):
    data = Incoming.objects.values('owner_fk__name_owner').annotate(part_count=Count('part_number'))
    labels = [entry['owner_fk__name_owner'] for entry in data]
    values = [entry['part_count'] for entry in data]
    return JsonResponse({'labels': labels, 'values': values}, safe=False)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def top_10_lowest_saldo(request):
    data = Incoming.objects.values('part_number', 'saldo').order_by('saldo')[:10]
    part_numbers = [entry['part_number'] for entry in data]
    saldos = [entry['saldo'] for entry in data]
    return JsonResponse({'part_numbers': part_numbers, 'saldos': saldos})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def soon_to_expire_parts(request):
    today = datetime.now()
    three_months_later = today + timedelta(days=90)
    six_months_later = today + timedelta(days=180)
    
    data = Incoming.objects.filter(f_vencimiento__gte=three_months_later, f_vencimiento__lte=six_months_later).order_by('f_vencimiento')
    
    labels = [entry.f_vencimiento.strftime('%Y-%m-%d') for entry in data]
    part_numbers = [entry.part_number for entry in data]
    
    return JsonResponse({'labels': labels, 'part_numbers': part_numbers})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def monthly_weight_chart(request):
    data = Comat.objects.annotate(month=TruncMonth('f_stdf')).values('month').annotate(total_weight=Sum('peso')).order_by('month')
    
    labels = [entry['month'].strftime('%Y-%m') for entry in data]
    weights = [entry['total_weight'] for entry in data]
    
    return JsonResponse({'labels': labels, 'weights': weights})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def monthly_cif_chart(request):
    data = Comat.objects.annotate(month=TruncMonth('f_stdf')).values('month').annotate(total_cif=Sum('sum_cif')).order_by('month')
    
    labels = [entry['month'].strftime('%Y-%m') for entry in data]
    cif_values = [entry['total_cif'] for entry in data]
    
    return JsonResponse({'labels': labels, 'cif_values': cif_values})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def state_pie_chart(request):
    data = Comat.objects.values('estado_fk__estado').annotate(stdf_count=Count('stdf_pk')).order_by('estado_fk__estado')
    
    labels = [entry['estado_fk__estado'] for entry in data]
    stdf_counts = [entry['stdf_count'] for entry in data]
    
    return JsonResponse({'labels': labels, 'stdf_counts': stdf_counts})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#BUSCADOR DE LA PAGINA DE INICIO, REDIRIGE AL RESULTADO DE BUSQUEDA DE LA PAGINA DE INICIO
def buscar_productos_inicio(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_inicio = request.GET.get('n', '')
    filtro = request.GET.get('f', '')
    return render(request, 'resultados_busqueda/resultado_busqueda_inicio.html', {'query_inicio':query_inicio, 'filtro': filtro})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def buscar_datos_inicio(request):
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    filtro = request.GET.get('f', '')
    search_value = request.GET.get('n', '')

    query_filter = Q()

    if filtro == '1':
        query_filter |= Q(stdf_fk__stdf_pk__icontains=search_value)
    elif filtro == '2':
        query_filter |= Q(part_number__icontains=search_value)

    query_incoming = Incoming.objects.filter(query_filter).select_related('stdf_fk', 'ubicacion_fk', 'categoria_fk', 'owner_fk').order_by('stdf_fk__stdf_pk')

    total_records_incoming = query_incoming.count()

    resultados_incoming_list = []
    
    for incoming in query_incoming:
        if incoming.categoria_fk.categoria_pk == 1:
            serial_number = incoming.sn_batch_pk
            batch_number = None
        elif incoming.categoria_fk.categoria_pk == 2:
            serial_number = None
            batch_number = incoming.sn_batch_pk
        elif incoming.categoria_fk.categoria_pk == 3:
            serial_number = incoming.sn_batch_pk
            batch_number = incoming.batch_pk
        else:
            serial_number = None
            batch_number = None


        qty_extraida_total = Consumos.objects.filter(incoming_fk=incoming.sn_batch_pk).aggregate(qty_extraida_total=Sum('qty_extraida'))['qty_extraida_total']
        
        resultados_incoming_list.append({
            "serial_number": serial_number,
            "batch_number": batch_number,
            "part_number": incoming.part_number,
            "descripcion": incoming.descripcion,
            "stdf_fk__stdf_pk": incoming.stdf_fk.stdf_pk,
            "qty": incoming.qty,
            "saldo": incoming.saldo,
            "stdf_fk__awb": incoming.stdf_fk.awb,
            "stdf_fk__num_manifiesto": incoming.stdf_fk.num_manifiesto,
            "owner_fk__name_owner": incoming.owner_fk.name_owner,
            "ubicacion_fk__name_ubicacion": incoming.ubicacion_fk.name_ubicacion,
            "f_vencimiento": incoming.f_vencimiento,
            "qty_extraida_total": qty_extraida_total,
        })

    resultados_incoming = resultados_incoming_list[start:start + length]  # Ajustar la paginación aquí

    # Crear el diccionario principal "data"
    data = {
        "draw": draw,
        "recordsTotal": total_records_incoming,
        "recordsFiltered": total_records_incoming,
        "data_incoming": resultados_incoming,
        # "data_comats_sin_incoming": resultados_comats_sin_incoming,   
    }

    return JsonResponse(data)



# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE DE LA PAGINA DE INICIO
def detalle_inicio(request, sn_batch_pk):

    detalle_inicio = Incoming.objects.get(sn_batch_pk=sn_batch_pk)   

    comat_data = detalle_inicio.stdf_fk

    consumos_data = detalle_inicio.consumos_set.all()

    return render(request,'tablas_detalle/detalle_inicio.html' , {'detalle_inicio':detalle_inicio, 'comat_data' : comat_data, 'consumos_data':consumos_data })

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
@login_required
def comat(request):
    get_form_comat = Comat.objects.all()
    total_cif = 0

    if request.method == 'POST':
        form_comat = ComatForm(request.POST)
        if form_comat.is_valid():
            if request.user.is_authenticated:
                # Obtén el objeto Comat sin guardarlo aún
                comat = form_comat.save(commit=False)
                comat.usuario = request.user  

                # Realiza la suma de fob, flete y seguro
                total_cif = comat.fob + comat.flete + comat.seguro

                # Guarda el objeto Comat con la cif actualizada
                comat.sum_cif = total_cif
                comat.save()
                messages.success(request, "Se ha Añadido Correctamente")
                return redirect('/comat')
            else:
                # Manejo del caso en el que el usuario no está autenticado
                return HttpResponse("Debes iniciar sesión para realizar esta acción.")
    else:
        form_comat = ComatForm()

    context = {
        'form_comat': form_comat,
        'get_form_comat': get_form_comat,
    }
    return render(request, 'formularios/comat.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#BUSCADOR DE COMAT
def buscar_productos(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_comat = request.GET.get('c', '')
    filtro = request.GET.get('t', '')

    return render(request, 'resultados_busqueda/resultado_busqueda_stdf.html', {'query_comat':query_comat, 'filtro':filtro})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS RESULTADOS CON MÁS RELACION QUE TIENE LA BUSQUEDA
def obtener_datos_comat(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('c', '')  # Término de búsqueda
    filtro = request.GET.get('t', '')

    comat_data = Comat.objects.all()
    
    query_filter = Q()

    if filtro == '1':
        query_filter |= Q(stdf_pk__icontains=search_value)
    elif filtro == '2':
        query_filter |= Q(awb__icontains=search_value) 
    elif filtro == '3':
        query_filter |= Q(hawb__icontains=search_value)

    comat_data = comat_data.filter(query_filter)

    # Ordenar por stdf_pk
    comat_data = comat_data.order_by('stdf_pk')
    # Realizar la consulta teniendo en cuenta la paginación
    comat_data = comat_data[start:start + length]


    # Formatea los datos en un formato compatible con DataTables
    data = []
    for comat in comat_data:
        data.append({
            "stdf_pk": comat.stdf_pk,
            "awb": comat.awb,
            "hawb":comat.hawb,
            "num_manifiesto":comat.num_manifiesto,
            "sum_cif":comat.sum_cif,
            "bodega_fk":comat.bodega_fk.name_bodega,
            "usuario": comat.usuario.username,
            
        })
    
    if search_value:
        records_filtered = Comat.objects.filter(stdf_pk__icontains=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Comat.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Comat.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered  # Total de registros después del filtrado (puedes ajustar esto según tus necesidades)
    })

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE POR LA ID
def detalle_comat(request, stdf_pk):
    detalle_comat = Comat.objects.get(stdf_pk=stdf_pk)    
    return render(request,'tablas_detalle/detalle_comat.html' , {'detalle_comat':detalle_comat})


# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#VISTA DE INCOMING QUE VALIDA EL FORMULARIO Y LO GUARDA Y REDIRIGE A LA PAGINA DE INCOMING
from django.core.cache import cache

@login_required 
def incoming(request):
    get_form_incoming = Incoming.objects.all()
    total_unit_cost = 0
    if request.method == 'POST':
        form_incoming = IncomingForm(request.POST)
        if form_incoming.is_valid():
                if request.user.is_authenticated:
                    # Guarda el formulario de Incoming
                    incoming = form_incoming.save(commit=False)

                    incoming.usuario = request.user

                    total_unit_cost =  incoming.qty * incoming.u_purchase_cost 
                    # Copia el valor de cantidad_extraida a la columna saldo
                    incoming.total_u_purchase_cost = total_unit_cost
                    incoming.saldo = incoming.qty
                    incoming.save()
                    request.session['incoming_fk'] = incoming.sn_batch_pk
                    messages.success(request, "Se ha Añadido Correctamente")
                    return redirect('/detalle_form')
                else:
                # Manejo del caso en el que el usuario no está autenticado
                    return HttpResponse("Debes iniciar sesión para realizar esta acción.")
    else:
        form_incoming = IncomingForm()

    context = {
        'form_incoming': form_incoming,
        'get_form_incoming': get_form_incoming, 
    }
    return render(request, 'formularios/incoming.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#GUARDA EL VALOR QUE SE BUSCO Y REDIRIGE A LA PAGINA DE RESULTADOS
def buscar_productos_incoming(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_inco = request.GET.get('e', '')
    filtro = request.GET.get('t', '')
    
    return render(request, 'resultados_busqueda/resultado_busqueda_incoming.html', {'query_inco':query_inco, 'filtro':filtro})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS DATOS RELACIONADOS A LA BUSQUEDA 
def obtener_datos_incoming(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('e', '')  # Término de búsqueda
    filtro = request.GET.get('t', '')

    incoming_data = Incoming.objects.all()
     
    query_filter = Q()

    if filtro == '1':
        query_filter |= Q(part_number__icontains = search_value) 
    elif filtro == '2':
        query_filter |= Q(sn_batch_pk__icontains=search_value) |Q(batch_pk__icontains=search_value)
    elif filtro == '3':
        query_filter |= Q(stdf_fk__stdf_pk=search_value)



    incoming_data = incoming_data.filter(query_filter)

    # Ordenar por stdf_pk
    incoming_data = incoming_data.order_by('stdf_fk')
        
    # Realizar la consulta teniendo en cuenta la paginación
    incoming_data = incoming_data[start:start + length]


    # Formatea los datos en un formato compatible con DataTables
    data = []
    for incoming in incoming_data:
        if incoming.categoria_fk.categoria_pk == 1:
            serial_number = incoming.sn_batch_pk
            batch_number = None
        elif incoming.categoria_fk.categoria_pk == 2:
            serial_number = None
            batch_number = incoming.sn_batch_pk
        elif incoming.categoria_fk.categoria_pk == 3:
            serial_number = incoming.sn_batch_pk
            batch_number = incoming.batch_pk
        else:
            serial_number = None
            batch_number = None
        data.append({
            "serial_number": serial_number,
            "batch_number": batch_number,
            "part_number":incoming.part_number,
            "descripcion": incoming.descripcion,
            "qty":incoming.qty,
            "f_vencimiento":incoming.f_vencimiento,
            "usuario": incoming.usuario.username,
            "saldo":incoming.saldo,
        })
    
    if search_value:
        records_filtered = Incoming.objects.filter(part_number__icontains=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Incoming.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Incoming.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered  # Total de registros después del filtrado (puedes ajustar esto según tus necesidades)
    })

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def detalle_incoming(request, sn_batch_pk):
    detalle_incoming = Incoming.objects.get(sn_batch_pk=sn_batch_pk)
    return render(request, 'tablas_detalle/detalle_incoming.html', {'detalle_incoming': detalle_incoming})




# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #


# ## Test
# def obtener_datos_stdf_incoming(request):

#     term = request.GET.get('q', '')

#     stdf_data = Comat.objects.filter(stdf_pk__icontains=term).values('stdf_pk')[:20]

#     # stdf_data = Comat.objects.all().values('stdf_pk')
#     stdf_list = list(stdf_data)
    
#     # Convierte la lista de diccionarios a una lista de objetos JSON
#     stdf_json = [{'stdf_pk': item['stdf_pk']} for item in stdf_list]
    
#     return JsonResponse({'stdf_data': stdf_json}, safe=False)

def obtener_datos_stdf_incoming(request):
    term = request.GET.get('q', '')

    # Filtra los objetos Comat según el término de búsqueda
    stdf_data = Comat.objects.filter(stdf_pk__icontains=term).first()

    # Verifica si se encontró un objeto Comat
    if stdf_data:
        # Convierte el objeto Comat a un diccionario
        comat_data = {
            'stdf_pk': stdf_data.stdf_pk,
            # Agrega otros campos de Comat que necesites
        }
        return JsonResponse({'stdf_data': comat_data}, safe=False)
    else:
        return JsonResponse({'stdf_data': None}, safe=False)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#VISTA DE CONSUMO QUE VALIDA EL FORMULARIO Y LO GUARDA Y REDIRIGE A LA VISTA DE CONSUMOS

def consumos(request):
    form_consumos = ConsumosForm()
    if request.method == 'POST':
        form_consumos = ConsumosForm(request.POST)
        if form_consumos.is_valid():
            if request.user.is_authenticated:
                consumo = form_consumos.save(commit=False)  # Guardar el formulario de Consumos sin guardarlo en la base de datos

                consumo.usuario = request.user

                # Obtener el registro de Incoming relacionado a través de ForeignKey
                incoming = consumo.incoming_fk  # Asegúrate de usar el nombre correcto del campo

                if consumo.qty_extraida <= incoming.saldo:
                    incoming.saldo -= consumo.qty_extraida
                    incoming.save()

                    # Guardar el registro de Consumos en la base de datos
                    consumo.save()
                    messages.success(request, "Se ha añadido correctamente")
                    return redirect('/consumos')
                else:
                    error_message = f"No puedes extraer más de lo que hay en el saldo. Saldo actual: {incoming.saldo}."
                    messages.error(request, error_message)
            else:
                # Manejo del caso en el que el usuario no está autenticado
                return HttpResponse("Debes iniciar sesión para realizar esta acción.")

    context = {
        'form_consumos': form_consumos,
    }
    return render(request, 'formularios/consumos.html', context)




# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#GUARDA LA BUSQUEDA QUE SE REALIZO Y REDIRIGE A LA PAGINA DE RESULTADOS DE CONSUMOS
def buscar_productos_consumos(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_consu = request.GET.get('t', '')

    return render(request, 'resultados_busqueda/resultado_busqueda_consumos.html', {'query_consu': query_consu})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS DATOS RELACIONADOS A LA BUSQUEDA DE CONSUMOS
def obtener_datos_consumos(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('t', '')  # Término de búsqueda

    # Inicializa la consulta sin filtrar
    consumos_data = Consumos.objects.all()

    # Realiza la consulta teniendo en cuenta el término de búsqueda en incoming_fk
    if search_value:
            consumos_data = consumos_data.filter(Q(incoming_fk=search_value))

    # Realiza la paginación
    consumos_data = consumos_data[start:start + length]

    # Formatea los datos en un formato compatible con DataTables
    data = []
    for consumo in consumos_data:
        data.append({
            "consumo_pk": consumo.consumo_pk,
            "incoming_fk": consumo.incoming_fk.sn_batch_pk,
            "f_transaccion": consumo.f_transaccion,
            "orden_consumo": consumo.orden_consumo,
            "matricula_aeronave": consumo.matricula_aeronave,
            "qty_extraida": consumo.qty_extraida,
            "usuario": consumo.usuario.username,
        })

    if search_value:
        records_filtered = Consumos.objects.filter(incoming_fk=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Consumos.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Consumos.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered,
        
    })

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE DE CONSUMOS
def detalle_consumos(request, consumo_pk):
    detalle_consumos = Consumos.objects.get(consumo_pk=consumo_pk)   

    incoming_data = detalle_consumos.incoming_fk

    return render(request,'tablas_detalle/detalle_consumos.html' , {'detalle_consumos':detalle_consumos, 'incoming_data' : incoming_data })


# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def obtener_datos_sn_consumos(request):
    term = request.GET.get('q', '')

    # Filtra los objetos Incoming según el término de búsqueda
    sn_data = Incoming.objects.filter(sn_batch_pk__icontains=term)[:10]

    print(f'Término de búsqueda: {term}')
    print(f'Número de resultados: {len(sn_data)}')

    # Verifica si se encontraron objetos Incoming
    if sn_data:
        # Convierte los objetos Incoming a una lista de diccionarios
        incoming_data = [
            {
                'sn_batch_pk': f"SN:{item.sn_batch_pk} - STDF:{item.stdf_fk.stdf_pk}",
        # Agrega otros campos de Incoming que necesites
            }
            for item in sn_data
        ]
        print(f'Datos de salida: {incoming_data}')
        return JsonResponse({'sn_data': incoming_data}, safe=False)
    else:
        # Retorna un diccionario vacío si no se encuentran resultados
        print('No se encontraron resultados')
        return JsonResponse({'sn_data': []}, safe=False)
# def obtener_datos_sn_consumos(request):

#     term = request.GET.get('q', '')

#     sn_data = Incoming.objects.filter(sn_batch_pk__icontains=term).values('sn_batch_pk')[:20]

#     sn_data = Incoming.objects.all().values('sn_batch_pk')
#     sn_list = list(sn_data)
    
#     # Convierte la lista de diccionarios a una lista de objetos JSON
#     sn_json = [{'sn_batch_pk': item['sn_batch_pk']} for item in sn_list]
    
#     return JsonResponse({'sn_data': sn_json}, safe=False)
    
# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def detalle_inicio(request, stdf_pk):
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página

    comat_data = Comat.objects.get(stdf_pk=stdf_pk)

    # Tabla de Incoming
    incoming_objects = Incoming.objects.filter(stdf_fk=comat_data)
    incoming_objects_paginated = incoming_objects[start:start + length]

    incoming_data_list = []

    for incoming_obj in incoming_objects_paginated:
        incoming_data_list.append({
            "sn_batch_pk": incoming_obj.sn_batch_pk,
            "part_number": incoming_obj.part_number,
            "f_incoming": incoming_obj.f_incoming,
            "descripcion": incoming_obj.descripcion,
            "po":incoming_obj.po,
            "qty":incoming_obj.qty,
            "u_purchase_cost":incoming_obj.u_purchase_cost,
            "total_u_purchase_cost":incoming_obj.total_u_purchase_cost,
            "f_vencimiento":incoming_obj.f_vencimiento,
            "saldo":incoming_obj.saldo,
            "observaciones":incoming_obj.observaciones,
            "categoria_fk":incoming_obj.categoria_fk.name_categoria,
            "clasificacion_fk":incoming_obj.clasificacion_fk.name_clasificacion,
            "ubicacion_fk":incoming_obj.ubicacion_fk.name_ubicacion,
            "uom_fk":incoming_obj.uom_fk.name_uom,
            "owner_fk":incoming_obj.owner_fk.name_owner,
            "condicion_fk":incoming_obj.condicion_fk.name_condicion,
            "ficha_fk":incoming_obj.ficha_fk.name_ficha,
    })

    incoming_records_total = incoming_objects.count()
    incoming_records_filtered = incoming_records_total  # Opcionalmente, puedes aplicar filtros aquí

    # Tabla de Consumos
    consumos_objects = Consumos.objects.filter(incoming_fk__in=incoming_objects)
    consumos_objects_paginated = consumos_objects[start:start + length]

    consumos_data_list = []

    for consumos_obj in consumos_objects_paginated:
        consumos_data_list.append({
                "incoming_fk":consumos_obj.incoming_fk.sn_batch_pk,
                "f_transaccion":consumos_obj.f_transaccion,
                "qty_extraida":consumos_obj.qty_extraida,
                "matricula_aeronave":consumos_obj.matricula_aeronave,
                "observaciones":consumos_obj.observaciones,
        })

    consumos_records_total = consumos_objects.count()
    consumos_records_filtered = consumos_records_total  # Opcionalmente, puedes aplicar filtros aquí

    data = {
        "draw": draw,
        "recordsTotal": incoming_records_total,
        "recordsFiltered": incoming_records_filtered,
        "comat_data": {
                "stdf_pk": comat_data.stdf_pk,
                "awb": comat_data.awb,
                "hawb": comat_data.hawb,
                "num_manifiesto": comat_data.num_manifiesto,
                "sum_cif": comat_data.sum_cif,
                "corr_interno" : comat_data.corr_interno,
                "psc": comat_data.pcs,
                "peso": comat_data.peso,
                "f_control": comat_data.f_control,
                "f_manifiesto": comat_data.f_manifiesto,
                "f_recepcion": comat_data.f_recepcion,
                "f_stdf": comat_data.f_stdf,
                "fob": comat_data.fob,
                "flete": comat_data.flete,
                "seguro": comat_data.seguro,
                "bodega_fk": comat_data.bodega_fk.name_bodega,
                "origen_fk": comat_data.origen_fk.name_origen,
                "estado_fk": comat_data.estado_fk.estado,
            },
        "incoming_data": incoming_data_list,
        "consumos_data": consumos_data_list,
    }

    if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
        return JsonResponse(data)
    else:
        # Si no es una solicitud AJAX, renderiza una plantilla HTML
        return render(request, 'tablas_detalle/detalle_inicio.html', {
            "comat_data": comat_data,
            "incoming_data_list": incoming_data_list,
            "incoming_recordsTotal": incoming_records_total,
            "incoming_recordsFiltered": incoming_records_filtered,
            "consumos_data": consumos_data_list,
            "consumos_recordsTotal": consumos_records_total,
            "consumos_recordsFiltered": consumos_records_filtered,
        })
    
# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def detalle_form(request):
    form1 = DetalleForm(prefix='form1')
    if request.method == 'POST':
        form1 = DetalleForm(request.POST, prefix='form1')
        if form1.is_valid():
            datos_form1 = form1.cleaned_data
            # Crea una instancia de Detalle_Incoming y guárdala
            modelo1 = Detalle_Incoming(**datos_form1)
            modelo1.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect(f'/detalle_incoming/{form1.cleaned_data["incoming_fk"]}/')

    return render(request, 'tablas_detalle/detalle_incomingforms.html', {'form1': form1})
# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

##################################
### Vista de Mantenedor Comat ####
##################################

def editar_comat(request, stdf_pk):
    comat = Comat.objects.get(pk=stdf_pk)

    if request.method == 'POST':
        form = ComatForm(request.POST, instance=comat)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/detalle_comat/'+str(stdf_pk))  # Redirige a la página deseada después de la edición.
    else:
        form = ComatForm(instance=comat)

    return render(request, 'formularios/editar_comat.html', {'form': form, 'comat': comat})


def eliminar_comat(request, stdf_pk):
    comats = Comat.objects.get(pk=stdf_pk)
    comats.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/buscar')


##################################
# Vista de Mantenedor Incoming   #
##################################

def editar_incoming(request, sn_batch_pk):
    incoming = Incoming.objects.get(pk=sn_batch_pk)

    if request.method == 'POST':
        form = IncomingForm(request.POST, instance=incoming)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/detalle_incoming/'+str(sn_batch_pk))  # Redirige a la página deseada después de la edición.
    else:
        form = IncomingForm(instance=incoming)

    return render(request, 'formularios/editar_incoming.html', {'form': form, 'incoming': incoming})


def eliminar_incoming(request, sn_batch_pk):
    incomings = Incoming.objects.get(pk=sn_batch_pk)
    incomings.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/buscar_incoming')

##################################
# Vista de Mantenedor Consumo   #
##################################

def editar_consumo(request, consumo_pk):
    try:
        consumo = Consumos.objects.get(consumo_pk=consumo_pk)
    except Consumos.DoesNotExist:
        return render(request, 'error.html', {'mensaje': 'El Consumo no existe'})

    if request.method == 'POST':
        # Guarda el valor actual de qty_extraida antes de actualizar el formulario
        qty_inicial = consumo.incoming_fk.qty
        qty_anterior = consumo.qty_extraida

        form = ConsumosForm(request.POST, instance=consumo)
        if form.is_valid():
            # Obtiene la cantidad total de cantidad extraída asociada al incoming
            total_cantidad_extraida = consumo.incoming_fk.total_cantidad_extraida()

            nuevo_qty_extraida = form.cleaned_data['qty_extraida']

            # Calcula la diferencia entre el valor anterior y el nuevo valor
            saldo = qty_inicial - total_cantidad_extraida + nuevo_qty_extraida

            consumo.incoming_fk.saldo = saldo
            incoming.save()  # Guarda los cambios en el objeto Incoming

            # Guarda los cambios en el objeto Consumo
            form.save()

            messages.success(request, "Se ha modificado correctamente.")
            return redirect('/detalle_consumos/' + str(consumo_pk))
    else:
        form = ConsumosForm(instance=consumo)

    return render(request, 'formularios/editar_consumo.html', {'form': form, 'consumo': consumo})

def eliminar_consumo(request, consumo_pk):
    consumos = Consumos.objects.get(consumo_pk=consumo_pk)
    incoming = consumos.incoming_fk
    incoming.saldo += consumos.qty_extraida
    incoming.save()
    consumos.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/buscar_consumos')

###########################################
### Vista de Mantenedores all ####
###########################################

def mantenedores_all(request):

    return render(request, 'mantenedores/mantenedores_all.html')

###########################################
### Vista de Mantenedor Categotia_incoming ####
###########################################

def mantenedor_categoria_incoming(request):
    get_categoria_incoming = Categotia_incoming.objects.all()

    context = {
        'get_categoria_incoming': get_categoria_incoming,
    }
    return render(request, 'mantenedores/categoria_incoming/mantenedor_categoria_incoming.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_categoria_incoming(request, categoria_pk):
    categoria_incoming = Categotia_incoming.objects.get(categoria_pk=categoria_pk)

    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria_incoming)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_categoria_incoming/')  # Redirige a la página deseada después de la edición.
    else:
        form = CategoriaForm(instance=categoria_incoming)

    return render(request, 'mantenedores/categoria_incoming/editar_categoria_incoming.html', {'form': form, 'categoria_incoming': categoria_incoming})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def registrar_categoria_incoming(request):
    form_reg_categoria = CategoriaForm()

    if request.method == 'POST':
        form_reg_categoria = CategoriaForm(request.POST)
        if form_reg_categoria.is_valid():
            if Categotia_incoming.objects.exists():
                last_register = Categotia_incoming.objects.latest('categoria_pk')
                new_pk = last_register.categoria_pk + 1
            else:
                new_pk = 1
            
            form_reg_categoria.instance.categoria_pk = new_pk
            form_reg_categoria.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_categoria_incoming')
        
    context = {
        'form_reg_categoria':form_reg_categoria
    }
        
    return render(request, 'mantenedores/categoria_incoming/registrar_categoria_incoming.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
def eliminar_categoria_incoming(request, categoria_pk):
    categoria_incoming = Categotia_incoming.objects.get(categoria_pk=categoria_pk)
    categoria_incoming.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_categoria_incoming')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor Estado         ####
###########################################

def mantenedor_estado(request):
    get_estado = Estado.objects.all()

    context = {
        'get_estado': get_estado,
    }
    return render(request, 'mantenedores/estado/mantenedor_estado.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_estado(request, estado_pk):
    estados = Estado.objects.get(estado_pk=estado_pk)

    if request.method == 'POST':
        form = EstadoForm(request.POST, instance=estados)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Moficado Correctamente")
            return redirect('/mantenedor_estado/')  # Redirige a la página deseada después de la edición.
    else:
        form = EstadoForm(instance=estados)

    return render(request, 'mantenedores/estado/editar_estado.html', {'form': form, 'estados': estados})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_estado(request):
    form_reg_estado = EstadoForm()

    if request.method == 'POST':
        form_reg_estado = EstadoForm(request.POST)
        if form_reg_estado.is_valid():
            if Estado.objects.exists():
                last_register = Estado.objects.latest('estado_pk')
                new_pk = last_register.estado_pk + 1
            else:
                new_pk = 1
            
            form_reg_estado.instance.estado_pk = new_pk
            form_reg_estado.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_estado')
        
    context = {
        'form_reg_estado':form_reg_estado
    }
        
    return render(request, 'mantenedores/estado/registrar_estado.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_estado(request, estado_pk):
    estados = Estado.objects.get(pk=estado_pk)
    estados.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_estado')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
###########################################
### Vista de Mantenedor Ubicación         ####
###########################################

def mantenedor_ubicacion(request):
    get_ubicacion = Ubicacion.objects.all()

    context = {
        'get_ubicacion': get_ubicacion,
    }
    return render(request, 'mantenedores/ubicacion/mantenedor_ubicacion.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_ubicacion(request, ubicacion_pk):
    ubicaciones = Ubicacion.objects.get(ubicacion_pk=ubicacion_pk)

    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicaciones)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_ubicacion/')  # Redirige a la página deseada después de la edición.
    else:
        form = UbicacionForm(instance=ubicaciones)

    return render(request, 'mantenedores/ubicacion/editar_ubicacion.html', {'form': form, 'ubicaciones': ubicaciones})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_ubicacion(request):
    form_reg_ubicacion = UbicacionForm()

    if request.method == 'POST':
        form_reg_ubicacion = UbicacionForm(request.POST)
        if form_reg_ubicacion.is_valid():
            if Ubicacion.objects.exists():
                last_register = Ubicacion.objects.latest('ubicacion_pk')
                new_pk = last_register.ubicacion_pk + 1
            else:
                new_pk = 1

            form_reg_ubicacion.instance.ubicacion_pk = new_pk
            form_reg_ubicacion.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_ubicacion')
        
    context = {
        'form_reg_ubicacion':form_reg_ubicacion
    }
        
    return render(request, 'mantenedores/ubicacion/registrar_ubicacion.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_ubicacion(request, ubicacion_pk):
    ubicaciones = Ubicacion.objects.get(pk=ubicacion_pk)
    ubicaciones.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_ubicacion')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor Uom   uom      ####
###########################################

def mantenedor_uom(request):
    get_uom = Uom.objects.all()

    context = {
        'get_uom': get_uom,
    }
    return render(request, 'mantenedores/uom/mantenedor_uom.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_uom(request, uom_pk):
    uoms = Uom.objects.get(uom_pk=uom_pk)

    if request.method == 'POST':
        form = UomForm(request.POST, instance=uoms)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_ubicacion/')  # Redirige a la página deseada después de la edición.
    else:
        form = UomForm(instance=uoms)

    return render(request, 'mantenedores/ubicacion/editar_ubicacion.html', {'form': form, 'uoms': uoms})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_uom(request):
    form_reg_uom = UomForm()

    if request.method == 'POST':
        form_reg_uom = UomForm(request.POST)
        if form_reg_uom.is_valid():
            if Uom.objects.exists():
                last_register = Uom.objects.latest('uom_pk')
                new_pk = last_register.uom_pk + 1
            else:
                new_pk = 1

            form_reg_uom.instance.uom_pk = new_pk
            form_reg_uom.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_uom')
        
    context = {
        'form_reg_uom':form_reg_uom
    }
        
    return render(request, 'mantenedores/uom/registrar_uom.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_uom(request, uom_pk):
    uoms = Uom.objects.get(pk=uom_pk)
    uoms.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_uom')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor owner     ####
###########################################

def mantenedor_owner(request):
    get_owner = Owner.objects.all()

    context = {
        'get_owner': get_owner,
    }
    return render(request, 'mantenedores/owner/mantenedor_owner.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_owner(request, owner_pk):
    owners = Owner.objects.get(owner_pk=owner_pk)

    if request.method == 'POST':
        form = OwnerForm(request.POST, instance=owners)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_owner/')  # Redirige a la página deseada después de la edición.
    else:
        form = OwnerForm(instance=owners)

    return render(request, 'mantenedores/owner/editar_owner.html', {'form': form, 'owners': owners})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_owner(request):
    form_reg_owner = OwnerForm()

    if request.method == 'POST':
        form_reg_owner = OwnerForm(request.POST)
        if form_reg_owner.is_valid():
            if Owner.objects.exists():
                last_register = Owner.objects.latest('owner_pk')
                new_pk = last_register.owner_pk + 1
            else:
                new_pk = 1

            form_reg_owner.instance.owner_pk = new_pk
            form_reg_owner.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_owner')
        
    context = {
        'form_reg_owner':form_reg_owner
    }
        
    return render(request, 'mantenedores/owner/registrar_owner.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_owner(request, owner_pk):
    owners = Owner.objects.get(pk=owner_pk)
    owners.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_owner')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor condition     ####
###########################################

def mantenedor_condition(request):
    get_condition = Condicion.objects.all()

    context = {
        'get_condition': get_condition,
    }
    return render(request, 'mantenedores/condition/mantenedor_condition.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_condition(request, condicion_pk):
    conditions = Condicion.objects.get(condicion_pk=condicion_pk)

    if request.method == 'POST':
        form = ConditionForm(request.POST, instance=conditions)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_condition/')  # Redirige a la página deseada después de la edición.
    else:
        form = ConditionForm(instance=conditions)

    return render(request, 'mantenedores/condition/editar_condition.html', {'form': form, 'conditions': conditions})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_condition(request):
    form_reg_condition = ConditionForm()

    if request.method == 'POST':
        form_reg_condition = ConditionForm(request.POST)
        if form_reg_condition.is_valid():
            if Condicion.objects.exists():
                last_register = Condicion.objects.latest('condicion_pk')
                new_pk = last_register.condicion_pk + 1
            else:
                new_pk = 1

            form_reg_condition.instance.condicion_pk = new_pk
            form_reg_condition.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_condition')
        
    context = {
        'form_reg_condition':form_reg_condition
    }
        
    return render(request, 'mantenedores/condition/registrar_condition.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_condition(request, condicion_pk):
    condicions = Condicion.objects.get(pk=condicion_pk)
    condicions.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_condition')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor ficha     ####
###########################################

def mantenedor_ficha(request):
    get_ficha = Ficha.objects.all()

    context = {
        'get_ficha': get_ficha,
    }
    return render(request, 'mantenedores/ficha/mantenedor_ficha.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_ficha(request, ficha_pk):
    fichas = Ficha.objects.get(ficha_pk=ficha_pk)

    if request.method == 'POST':
        form = FichaForm(request.POST, instance=fichas)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_ficha/')  # Redirige a la página deseada después de la edición.
    else:
        form = FichaForm(instance=fichas)

    return render(request, 'mantenedores/ficha/editar_ficha.html', {'form': form, 'fichas': fichas})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_ficha(request):
    form_reg_ficha = FichaForm()

    if request.method == 'POST':
        form_reg_ficha = FichaForm(request.POST)
        if form_reg_ficha.is_valid():
            if Ficha.objects.exists():
                last_register = Ficha.objects.latest('ficha_pk')
                new_pk = last_register.ficha_pk + 1
            else:
                new_pk = 1

            form_reg_ficha.instance.ficha_pk = new_pk
            form_reg_ficha.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_ficha')
        
    context = {
        'form_reg_ficha':form_reg_ficha
    }
        
    return render(request, 'mantenedores/ficha/registrar_ficha.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_ficha(request, ficha_pk):
    fichas = Ficha.objects.get(pk=ficha_pk)
    fichas.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_ficha')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor bodega     ####
###########################################

def mantenedor_bodega(request):
    get_bodega = Bodega.objects.all()

    context = {
        'get_bodega': get_bodega,
    }
    return render(request, 'mantenedores/bodega/mantenedor_bodega.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_bodega(request, bodega_pk):
    bodegas = Bodega.objects.get(bodega_pk=bodega_pk)

    if request.method == 'POST':
        form = BodegaForm(request.POST, instance=bodegas)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_bodega/')  # Redirige a la página deseada después de la edición.
    else:
        form = BodegaForm(instance=bodegas)

    return render(request, 'mantenedores/bodega/editar_bodega.html', {'form': form, 'bodegas': bodegas})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_bodega(request):
    form_reg_bodega = BodegaForm()

    if request.method == 'POST':
        form_reg_bodega = BodegaForm(request.POST)
        if form_reg_bodega.is_valid():
            if Bodega.objects.exists():
                last_register = Bodega.objects.latest('bodega_pk')
                new_pk = last_register.bodega_pk + 1
            else:
                new_pk = 1

            form_reg_bodega.instance.bodega_pk = new_pk
            form_reg_bodega.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_bodega')
        
    context = {
        'form_reg_bodega':form_reg_bodega
    }
        
    return render(request, 'mantenedores/bodega/registrar_bodega.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_bodega(request, bodega_pk):
    bodegas = Bodega.objects.get(pk=bodega_pk)
    bodegas.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_bodega')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor origen     ####
###########################################

def mantenedor_origen(request):
    get_origen = Origen.objects.all()

    context = {
        'get_origen': get_origen,
    }
    return render(request, 'mantenedores/origen/mantenedor_origen.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_origen(request, origen_pk):
    origens = Origen.objects.get(origen_pk=origen_pk)

    if request.method == 'POST':
        form = OrigenForm(request.POST, instance=origens)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_origen/')  # Redirige a la página deseada después de la edición.
    else:
        form = OrigenForm(instance=origens)

    return render(request, 'mantenedores/origen/editar_origen.html', {'form': form, 'origens': origens})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_origen(request):
    form_reg_origen = OrigenForm()

    if request.method == 'POST':
        form_reg_origen = OrigenForm(request.POST)
        if form_reg_origen.is_valid():
            if Origen.objects.exists():
                last_register = Origen.objects.latest('origen_pk')
                new_pk = last_register.origen_pk + 1
            else:
                new_pk = 1

            form_reg_origen.instance.origen_pk = new_pk
            form_reg_origen.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_origen')
        
    context = {
        'form_reg_origen':form_reg_origen
    }
        
    return render(request, 'mantenedores/origen/registrar_origen.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_origen(request, origen_pk):
    origenes = Origen.objects.get(pk=origen_pk)
    origenes.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_origen')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor cargo     ####
###########################################

def mantenedor_cargo(request):
    get_cargo = Cargo.objects.all()

    context = {
        'get_cargo': get_cargo,
    }
    return render(request, 'mantenedores/cargo/mantenedor_cargo.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_cargo(request, id):
    cargos = Cargo.objects.get(id=id)

    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargos)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_cargo/')  # Redirige a la página deseada después de la edición.
    else:
        form = CargoForm(instance=cargos)

    return render(request, 'mantenedores/cargo/editar_cargo.html', {'form': form, 'cargos': cargos})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_cargo(request):
    form_reg_cargo = CargoForm()

    if request.method == 'POST':
        form_reg_cargo = CargoForm(request.POST)
        if form_reg_cargo.is_valid():
            if Cargo.objects.exists():
                last_register = Cargo.objects.latest('id')
                new_pk = last_register.id + 1
            else:
                new_pk = 1

            form_reg_cargo.instance.id = new_pk
            form_reg_cargo.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_cargo')
        
    context = {
        'form_reg_cargo':form_reg_cargo
    }
        
    return render(request, 'mantenedores/cargo/registrar_cargo.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_cargo(request, id):
    cargos = Cargo.objects.get(pk=id)
    cargos.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_cargo')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor clasificacion     ####
###########################################

def mantenedor_clasificacion(request):
    get_clasificacion = Clasificacion.objects.all()

    context = {
        'get_clasificacion': get_clasificacion,
    }
    return render(request, 'mantenedores/clasificacion/mantenedor_clasificacion.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_clasificacion(request, clasificacion_pk):
    clasificaciones = Clasificacion.objects.get(clasificacion_pk=clasificacion_pk)

    if request.method == 'POST':
        form = ClasificacionForm(request.POST, instance=clasificaciones)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_clasificacion/')  # Redirige a la página deseada después de la edición.
    else:
        form = ClasificacionForm(instance=clasificaciones)

    return render(request, 'mantenedores/clasificacion/editar_clasificacion.html', {'form': form, 'clasificaciones': clasificaciones})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_clasificacion(request):
    form_reg_clasificacion = ClasificacionForm()

    if request.method == 'POST':
        form_reg_clasificacion = ClasificacionForm(request.POST)
        if form_reg_clasificacion.is_valid():
            if Clasificacion.objects.exists():
                last_register = Clasificacion.objects.latest('clasificacion_pk')
                new_pk = last_register.clasificacion_pk + 1
            else:
                new_pk = 1

            form_reg_clasificacion.instance.clasificacion_pk = new_pk
            form_reg_clasificacion.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_clasificacion')
        
    context = {
        'form_reg_clasificacion':form_reg_clasificacion
    }
        
    return render(request, 'mantenedores/clasificacion/registrar_clasificacion.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_clasificacion(request, clasificacion_pk):
    clasificaciones = Clasificacion.objects.get(pk=clasificacion_pk)
    clasificaciones.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_clasificacion')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #
###########################################
### Vista de Mantenedor compañia     ####
###########################################

def mantenedor_compañia(request):
    get_compañia = Compania.objects.all()

    context = {
        'get_compañia': get_compañia,
    }
    return render(request, 'mantenedores/compañia/mantenedor_compañia.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_compañia(request, cod_compania):
    compañias = Compania.objects.get(cod_compania=cod_compania)

    if request.method == 'POST':
        form = CompaniaForm(request.POST, instance=compañias)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_compañia/')  # Redirige a la página deseada después de la edición.
    else:
        form = CompaniaForm(instance=compañias)

    return render(request, 'mantenedores/compañia/editar_compañia.html', {'form': form, 'compañias': compañias})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_compañia(request):
    form_reg_compañia = CompaniaForm()

    if request.method == 'POST':
        form_reg_compañia = CompaniaForm(request.POST)
        if form_reg_compañia.is_valid():
            if Compania.objects.exists():
                last_register = Compania.objects.latest('cod_compania')
                new_pk = last_register.cod_compania + 1
            else:
                new_pk = 1

            form_reg_compañia.instance.cod_compania = new_pk
            form_reg_compañia.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_compañia')
        
    context = {
        'form_reg_compañia':form_reg_compañia
    }
        
    return render(request, 'mantenedores/compañia/registrar_compañia.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_compañia(request, cod_compania):
    companias = Compania.objects.get(pk=cod_compania)
    companias.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_compañia')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor consumidor     ####
###########################################

def mantenedor_consumidor(request):
    get_consumidor = Consumidor.objects.all()

    context = {
        'get_consumidor': get_consumidor,
    }
    return render(request, 'mantenedores/consumidor/mantenedor_consumidor.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_consumidor(request, id):
    consumidors = Consumidor.objects.get(id=id)

    if request.method == 'POST':
        form = ConsumidorForm(request.POST, instance=consumidors)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_consumidor/')  # Redirige a la página deseada después de la edición.
    else:
        form = ConsumidorForm(instance=consumidors)

    return render(request, 'mantenedores/consumidor/editar_consumidor.html', {'form': form, 'consumidors': consumidors})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_consumidor(request):
    form_reg_consumidor = ConsumidorForm()

    if request.method == 'POST':
        form_reg_consumidor = ConsumidorForm(request.POST)
        if form_reg_consumidor.is_valid():
            if Consumidor.objects.exists():
                last_register = Consumidor.objects.latest('id')
                new_pk = last_register.id + 1
            else:
                new_pk = 1

            form_reg_consumidor.instance.id = new_pk
            form_reg_consumidor.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_consumidor')
        
    context = {
        'form_reg_consumidor':form_reg_consumidor
    }
        
    return render(request, 'mantenedores/consumidor/registrar_consumidor.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_consumidor(request, id):
    consumidores = Consumidor.objects.get(pk=id)
    consumidores.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_consumidor')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

###########################################
### Vista de Mantenedor estado_repuesto     ####
###########################################

def mantenedor_estado_repuesto(request):
    get_estado_repuesto = Estado_Repuesto.objects.all()

    context = {
        'get_estado_repuesto': get_estado_repuesto,
    }
    return render(request, 'mantenedores/estado_repuesto/mantenedor_estado_repuesto.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_estado_repuesto(request, id):
    estado_repuestos = Estado_Repuesto.objects.get(id=id)

    if request.method == 'POST':
        form = EstadoRepuestoForm(request.POST, instance=estado_repuestos)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_estado_repuesto/')  # Redirige a la página deseada después de la edición.
    else:
        form = EstadoRepuestoForm(instance=estado_repuestos)

    return render(request, 'mantenedores/estado_repuesto/editar_estado_repuesto.html', {'form': form, 'estado_repuestos': estado_repuestos})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_estado_repuesto(request):
    form_reg_estado_repuesto = EstadoRepuestoForm()

    if request.method == 'POST':
        form_reg_estado_repuesto = EstadoRepuestoForm(request.POST)
        if form_reg_estado_repuesto.is_valid():
            if Estado_Repuesto.objects.exists():
                last_register = Estado_Repuesto.objects.latest('id')
                new_pk = last_register.id + 1
            else:
                new_pk = 1

            form_reg_estado_repuesto.instance.id = new_pk
            form_reg_estado_repuesto.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_estado_repuesto')
        
    context = {
        'form_reg_estado_repuesto':form_reg_estado_repuesto
    }
        
    return render(request, 'mantenedores/estado_repuesto/registrar_estado_repuesto.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_estado_repuesto(request, id):
    estado_repuestos = Estado_Repuesto.objects.get(pk=id)
    estado_repuestos.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_estado_repuesto')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

##################################
### Vista de Mantenedor DetalleForm ####
##################################

def editar_detalle_incoming_form(request, sn_batch_pk):
    detalleForm = Detalle_Incoming.objects.get(incoming_fk=sn_batch_pk)

    if request.method == 'POST':
        form = DetalleForm(request.POST, instance=detalleForm)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/detalle_incoming/'+sn_batch_pk)  # Redirige a la página deseada después de la edición.
    else:
        form = DetalleForm(instance=detalleForm)

    return render(request, 'mantenedores/detalle_incoming_form/editar_detalle_incomingforms.html', {'form': form, 'detalleForm': detalleForm})

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #


##################################
### Vista de Mantenedor Licencia ####
##################################

def mantenedor_licencia(request):
    get_licencia = Licencia.objects.all()

    context = {
        'get_licencia': get_licencia,
    }
    return render(request, 'mantenedores/licencia/mantenedor_licencia.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def editar_licencia(request, id):
    licencias = Licencia.objects.get(id=id)

    if request.method == 'POST':
        form = LicenciaForm(request.POST, instance=licencias)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha Modificado Correctamente")
            return redirect('/mantenedor_licencia/')  # Redirige a la página deseada después de la edición.
    else:
        form = LicenciaForm(instance=licencias)

    return render(request, 'mantenedores/licencia/editar_licencia.html', {'form': form, 'licencias': licencias})


# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def registrar_licencia(request):
    form_reg_licencia = LicenciaForm()

    if request.method == 'POST':
        form_reg_licencia = LicenciaForm(request.POST)
        if form_reg_licencia.is_valid():
            if Licencia.objects.exists():
                last_register = Licencia.objects.latest('id')
                new_pk = last_register.id + 1
            else:
                new_pk = 1

            form_reg_licencia.instance.id = new_pk
            form_reg_licencia.save()
            messages.success(request, "Se ha Añadido Correctamente")
            return redirect('/mantenedor_licencia')
        
    context = {
        'form_reg_licencia':form_reg_licencia
    }
        
    return render(request, 'mantenedores/licencia/registrar_licencia.html', context)

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def eliminar_licencia(request, id):
    licencias = Licencia.objects.get(pk=id)
    licencias.delete()
    messages.success(request, "Se ha Eliminado Correctamente")
    return redirect('/mantenedor_licencia')

# -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- ## -- # -- # -- # -- # -- # -- #

def estadostdf(request):
    with connection.cursor() as cursor:
        cursor.execute('CALL abona_cancela()')
    messages.success(request, "Se ha ejecutado Correctamente")
    return redirect('/orden_consumo')



def orden_consumo(request):
    if request.method == 'POST':
        form = OrdenConsumoForm(request.POST)
        if form.is_valid():
            # Procesar las fechas ingresadas, por ejemplo, guardar en la base de datos o realizar cálculos
            fechainicio = form.cleaned_data['fechainicio']
            fechatermino = form.cleaned_data['fechatermino']
            # Agregar aquí la lógica de procesamiento

            consumos = Consumos.objects.filter(f_transaccion__range=(fechainicio,fechatermino))

            wk = openpyxl.Workbook()
            ws = wk.active

            # Agregar encabezados
            ws['A1'] = 'Fecha de Transacción'
            ws['A2'] = 'Descripcion'
            ws['A3'] = 'STDF'
            ws['A4'] = 'Cancela o Abona'
            # Agregar otros encabezados según tus campos

            # Agregar datos de consumos
            row = 2
            for consumo in consumos:
                ws.cell(row=row, column=1, value=consumo.f_transaccion)
                ws.cell(row=row, column=2, value=consumo.incoming_fk.descripcion)
                ws.cell(row=row, column=3, value=consumo.incoming_fk.stdf_fk)
                ws.cell(row=row, column=4, value=consumo.incoming_fk.stdf_fk.estado_fk)
                # Agregar otros campos de consumo según tus campos
                row += 1

            # Guardar el archivo Excel en la memoria
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="consumos.xlsx"'
            wk.save(response)

            return response

    else:
        form = OrdenConsumoForm()

    return render(request, 'orden_consumo.html', {'form': form})