from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , Border, Side
from .forms import *
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
import locale
import datetime

def orden_consumos(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="orden.xlsx"'
    excel_file_path = 'staticfiles/excel/footer.xlsx'
    
    if request.method == 'POST':
        form = OrdenConsumoForm(request.POST)
        if form.is_valid():
            # Procesar las fechas ingresadas, por ejemplo, guardar en la base de datos o realizar cálculos
            fechainicio = form.cleaned_data['fechainicio']
            fechatermino = form.cleaned_data['fechatermino']
            # Agregar aquí la lógica de procesamiento

            compania = form.cleaned_data['compania']
            aduana = form.cleaned_data['aduana']
            resolucion_habilitacion = form.cleaned_data['resolucion_habilitacion']
            orden_consumo = form.cleaned_data['orden_consumo']

            consumos = Consumos.objects.filter(f_transaccion__range=(fechainicio,fechatermino),incoming_fk__stdf_fk__compania_fk__nom_compania=compania)

            for consumo in consumos:
                consumo.orden_consumo = orden_consumo
                consumo.save()

            wk2 = openpyxl.load_workbook(excel_file_path)
            ws2 = wk2.active

            wk = Workbook()
            ws = wk.active

            fecha_actual = datetime.datetime.now()

            # Formatear la fecha como una cadena de texto en el formato deseado (día/mes/año)
            fecha_formateada = fecha_actual.strftime("%d/%m/%Y")

            margen_izquierdo = 0.25
            margen_derecho = 0.25
            margen_superior = 0.6
            margen_inferior = 0.98

            # Configurar los márgenes de la hoja
            ws.page_margins.left = margen_izquierdo
            ws.page_margins.right = margen_derecho
            ws.page_margins.top = margen_superior
            ws.page_margins.bottom = margen_inferior

            row = 9  # Inicializa 'row' aquí

            
            # Configura el entorno local para usar comas como separador de miles y punto como separador decimal
            locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')  # Aquí se usa 'es_ES' como ejemplo de configuración regional en español
            # Formatea el número con "USD", comas para los miles y punto para los decimales
            
            
        
            # Resto de la lógica
            for consumo in consumos:
                u_purchase_cost_formateado = locale.format_string('%.2f', consumo.incoming_fk.u_purchase_cost * consumo.qty_extraida, grouping=True)

                ws.cell(row=row, column=1, value=consumo.qty_extraida)
                ws.cell(row=row, column=2, value=consumo.incoming_fk.ficha_fk.name_ficha)
                ws.cell(row=row, column=3, value='DEP. FRANCO')
                ws.cell(row=row, column=4, value='U-10')
                descripcion = f"{consumo.incoming_fk.descripcion} PN {consumo.incoming_fk.part_number} SN {consumo.incoming_fk.sn_batch_pk}"
                ws.cell(row=row, column=5, value=descripcion)
                ws.cell(row=row, column=6, value=consumo.matricula_aeronave)
                ws.cell(row=row, column=7, value=consumo.incoming_fk.stdf_fk.stdf_pk)
                ws.cell(row=row, column=8, value=consumo.incoming_fk.stdf_fk.estado_fk.estado)
                ws.cell(row=row, column=9, value=f"USD {u_purchase_cost_formateado}")
                row += 1

            ancho_columnas = {
                'A': 11.00,  
                'B': 6.89,  
                'C': 12.50,
                'D': 8.50,
                'E': 43.56,
                'F': 11.11,
                'G': 8.11,
                'H': 12.50,  
                'I': 15.56,
                
            }

            # Itera sobre las columnas y establece el ancho
            for columna, ancho in ancho_columnas.items():
                ws.column_dimensions[columna].width = ancho

            alto_filas = {
                3: 17.40,  # Alto de la fila 3
                8: 23.40,  # Alto de la fila 8
                
            }

            alto_deseado = 13.20

            # Iterar sobre las filas y cambiar el alto
            for fila in ws.iter_rows():
                ws.row_dimensions[fila[0].row].height = alto_deseado

                fuente = Font(size=9, bold=True)

            # Itera sobre las filas y cambia el alto según lo especificado en alto_filas
            for fila, alto in alto_filas.items():
                ws.row_dimensions[fila].height = alto

            ws.merge_cells('A1:C1')
            ws['A1'] = 'SERVICIO NACIONAL DE ADUANAS'

            ws.merge_cells('A3:F3')
            ws['A3'] = 'ORDEN DE CONSUMO'

            ws.merge_cells('A4:C4')
           

            ws.merge_cells('A5:C5')
            ws['A5'] = 'ADUANA DE PRESENTACION'
            ws['D5'] = ':'
            ws['E5'] = aduana


            ws.merge_cells('A6:C6')
            ws['A6'] = 'COMPAÑÍA QUE PRESENTA'
            ws['D6'] = ':'

            ws['E6'] = compania

            ws.merge_cells('A7:C7')
            ws['A7'] = 'RESOLUCION DE HABILITACION'
            ws['D7'] = ':'

            ws['E7'] = resolucion_habilitacion

            ws['H6'] = orden_consumo
            ws['I6'] = fecha_formateada

            ws['A8'] = 'CANTIDADES'

            ws['B8'] = 'FICHA'

            ws['C8'] = 'UBICACIÓN'

            ws['D8'] = 'CODIGO'

            ws['E8'] = 'DESCRIPCION DE MERCANCÍAS'

            ws['F8'] = 'MATRICULA AERONAVE'

            ws['G8'] = 'STDF'

            ws['H8'] = 'CANCELA O ABONA'

            ws['I8'] = 'VALOR FOB US$$'

           


            ws.merge_cells('G1:G7')# 4 BORDES
            ws.merge_cells('H1:I5')# SUPERIOR IZQUIERDO Y DERECHO

            borde_superior = Border(top=Side(style='thin'))
            borde_derecho = Border(right=Side(style='thin'))
            borde_inferior = Border(bottom=Side(style='thin'))
            borde_izquierdo = Border(left=Side(style='thin'))

            # Aplicar los bordes a las celdas específicas
            for columna in range(1, 10):  # Columnas A a I
                ws.cell(row=1, column=columna).border = borde_superior  # Borde superior

            for fila in range(1, 39):  # Filas 1 a 50
                ws.cell(row=fila, column=9).border = borde_derecho  # Borde derecho en la columna I
                ws.cell(row=fila, column=1).border = borde_izquierdo  # Borde izquierdo en la columna A

            borde_superior_derecho = Border(top=Side(style='thin'), right=Side(style='thin'))

            # Definir la celda a la que deseas aplicar los bordes
            celdaI1 = ws['I1']

            celdaI1.border = borde_superior_derecho

            borde_superior_izquierdo = Border(top=Side(style='thin'), left=Side(style='thin'))

            celdaA1 = ws['A1']

            celdaA1.border = borde_superior_izquierdo

            borde_4 = Border(top=Side(style='thin'), 
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            bottom=Side(style='thin'))

            celdas_G1G7 = ws['G1:G7']

            # Aplicar el borde a todas las celdas en el rango de celdas fusionadas
            for fila in celdas_G1G7:
                for celda in fila:
                    celda.border = borde_4

            borde_arriba_izquierdo_derecho = Border(top=Side(style='thin'), 
                            left=Side(style='thin'),
                            right=Side(style='thin'))
            celdas_H1I5 = ws['H1:I5']

            # Aplicar el borde a todas las celdas en el rango de celdas fusionadas
            for fila in celdas_H1I5:
                for celda in fila:
                    celda.border = borde_arriba_izquierdo_derecho

            borde_arriba_abajo_izquierda = Border(top=Side(style='thin'), 
                                        bottom=Side(style='thin'),
                                        left=Side(style='thin'))

            celdas_A5C5 = ws['A5':'C5']

            # Aplicar el borde a todas las celdas en el rango de celdas fusionadas
            for fila in celdas_A5C5:
                for celda in fila:
                    celda.border = borde_arriba_abajo_izquierda



            borde_arriba_abajo = Border(top=Side(style='thin'), 
                                        bottom=Side(style='thin'))

            celdaD5 = ws['D5']
            celdaE5 = ws['E5']

            # Aplicar el borde a la celda
            celdaD5.border = borde_arriba_abajo
            celdaE5.border = borde_arriba_abajo

            borde_arriba_abajo_derecho = Border(top=Side(style='thin'), 
                                                bottom=Side(style='thin'),
                                                right=Side(style='thin'))

            celdaF5 = ws['F5']


            # Aplicar el borde a la celda
            celdaF5.border = borde_arriba_abajo_derecho

            borde_arriba_izquierda = Border(top=Side(style='thin'), 
                                        left=Side(style='thin'),)

            celdas_A7C7 = ws['A7':'C7']

            # Aplicar el borde a todas las celdas en el rango de celdas fusionadas
            for fila in celdas_A7C7:
                for celda in fila:
                    celda.border = borde_arriba_izquierda

            borde_arriba= Border(top=Side(style='thin'))
            celdaD7 = ws['D7']
            celdaE7 = ws['E7']

            # Aplicar el borde a la celda
            celdaD7.border = borde_arriba
            celdaE7.border = borde_arriba

            borde_arriba_derecha= Border(top=Side(style='thin'),
                                        right=Side(style='thin'))
            celdaF7 = ws['F7']


            # Aplicar el borde a la celda
            celdaF7.border = borde_arriba_derecha


            borde_completo = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Iterar sobre las celdas en la fila A8 a la I8
            for columna in range(1, 10):  # Columnas A a I
                celda = ws.cell(row=8, column=columna)
                celda.border = borde_completo  # Aplicar los cuatro bordes a cada celda



            borde_lado = Border(
                left=Side(style='thin'),
                right=Side(style='thin')
            )

            # Iterar sobre las filas y columnas para aplicar los bordes
            for fila in range(9, 39):  # Filas de la 10 a la 38
                for columna in range(1, 10):  # Columnas A a I
                    celda = ws.cell(row=fila, column=columna)
                    celda.border = borde_lado  # Aplicar el

            ultima_fila = ws.max_row

            # Definir la fila a partir de la cual deseas aplicar los bordes (en este caso, 39)
            fila_inicio = 39

            # Iterar sobre las filas y columnas para aplicar los bordes
            for fila in range(fila_inicio, ultima_fila + 1):  # Desde la fila 39 hasta la última fila
                for columna in range(1, 10): 
                    celda = ws.cell(row=fila, column=columna)
                    celda.border = borde_lado

            alineacion_izquierda_ajuste = Alignment(horizontal='left', wrap_text=False)

            # Crear un objeto Font para definir el estilo de fuente (tamaño 8 y negrita)
            fuente = Font(name='Arial', size=10, bold=True)

            # Iterar sobre todas las celdas desde la fila 1 a la 7
            for fila in range(1, 8):  # Filas de la 1 a la 7
                for columna in range(1, ws.max_column + 1):  # Todas las columnas
                    celda = ws.cell(row=fila, column=columna)
                    celda.alignment = alineacion_izquierda_ajuste
                    celda.font = fuente


            alineacion_centro = Alignment(horizontal='center', vertical='center')
            fuente1 = Font(name='Arial', size=14, bold=True)

            # Aplicar la alineación al centro a todas las celdas en la fila 3
            for columna in range(1, 10):
                celda = ws.cell(row=3, column=columna)
                celda.alignment = alineacion_centro
                celda.font = fuente1

            alineacion_centro = Alignment(horizontal='center', vertical='center')
            fuente1 = Font(name='Arial', size=14, bold=True)

            # Aplicar la alineación al centro a todas las celdas en la fila 3
            for columna in range(1, 10):
                celda = ws.cell(row=3, column=columna)
                celda.alignment = alineacion_centro
                celda.font = fuente1

            alineacion_centro = Alignment(horizontal='left')
            fuente2 = Font(name='Arial', size=8, bold=True)

            # Aplicar la alineación al centro a todas las celdas en la fila 3
            for columna in range(1, 10):
                celda = ws.cell(row=1, column=columna)
                celda.alignment = alineacion_centro
                celda.font = fuente2



            alineacion_centro_8 = Alignment(horizontal='center', vertical='center',wrap_text=True)
            fuente3 = Font(name='tahoma', size=8, bold=True)

            # Aplicar la alineación al centro a todas las celdas en la fila 3
            for columna in range(1, 10):
                celda = ws.cell(row=8, column=columna)
                celda.alignment = alineacion_centro_8
                celda.font = fuente3


            alineacion_centro = Alignment(horizontal='center', vertical='center',wrap_text=True )

            # Crear un objeto Font para definir el estilo de fuente (Tahoma, tamaño 8)
            fuente = Font(name='Tahoma', size=8)

            # Encontrar la última fila con datos en la hoja
            ultima_fila = ws.max_row

            # Iterar sobre las filas desde la fila 8 hasta la última fila con datos
            for fila in range(9, ultima_fila + 1):
                for columna in range(1, ws.max_column + 1):
                    celda = ws.cell(row=fila, column=columna)
                    celda.alignment = alineacion_centro
                    celda.font = fuente


            ultima_fila = ws.max_row + 1

            # Copia el contenido de las 12 líneas de la hoja de origen a la hoja de destino
            for fila in range(ultima_fila, ultima_fila + 12):
                for columna in range(1, 10):
                    celda_origen = ws2.cell(row=fila - ultima_fila + 1, column=columna)
                    celda_destino = ws.cell(row=fila, column=columna, value=celda_origen.value)

                    # Copia los bordes de la celda de origen a la celda de destino
                    if celda_origen.border:
                        celda_destino.border = Border(
                            left=celda_origen.border.left,
                            right=celda_origen.border.right,
                            top=celda_origen.border.top,
                            bottom=celda_origen.border.bottom
                        )

                    # Copia la alineación de la celda de origen a la celda de destino
                    if celda_origen.alignment:
                        alineacion = Alignment(
                            horizontal=celda_origen.alignment.horizontal,
                            vertical=celda_origen.alignment.vertical,
                            text_rotation=celda_origen.alignment.text_rotation,
                            wrap_text=celda_origen.alignment.wrap_text,
                            shrink_to_fit=celda_origen.alignment.shrink_to_fit,
                            indent=celda_origen.alignment.indent
                        )
                        celda_destino.alignment = alineacion

            # Copia el formato de fuente (Arial 8) de la hoja de origen a la hoja de destino
            fuente_origen = ws2.cell(row=1, column=1).font  # Supongamos que el formato de fuente es el mismo en toda la hoja de origen
            fuente_destino = Font(name='Arial', size=8, bold=True)
            for fila in range(ultima_fila, ultima_fila + 12):
                for columna in range(1, 10):
                    celda_destino = ws.cell(row=fila, column=columna)
                    celda_destino.font = fuente_destino

            suma = 0

            # Itera sobre las filas desde la fila 9 hasta la última fila con datos
            for row in range(9, ws.max_row + 1):
                # Obtiene el valor de la celda en la columna 9
                celda = ws.cell(row=row, column=9).value
                # Si la celda comienza con "USD ", quita ese texto antes de convertir el valor a un número
                if celda and celda.startswith("USD "):
                    valor = celda.replace('USD ', '').replace(',', '.')
                    suma += float(valor)

            # Calcula la fila de destino restando 12 filas a la última fila con datos
            fila_destino = ws.max_row - 11

            # Establece la suma en la celda 12 filas antes de la última fila con datos en la columna 9
            ws.cell(row=fila_destino, column=9, value=f"USD {locale.format_string('%.2f', suma, grouping=True)}")




            
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = True
            ws.page_setup.fitToHeight = False

            texto_pie_pagina = "&C Orden de Consumos - Página &[Page] de &[Pages]"
            ws.oddFooter.left.text = texto_pie_pagina

            wk.save(response)
            return response
    else:
            form = OrdenConsumoForm()

            return render(request, 'orden_consumo.html', {'form': form})
