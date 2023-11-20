from .forms import *
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , Border, Side
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage


def exportar_excel_incoming(request, sn_batch_pk):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    imagen_path = 'staticfiles\\img\\Imagen1.png'
    image = Image.open(imagen_path)

    nuevo_ancho = 122
    nueva_altura = 97

# Redimensiona la imagen
    image = image.resize((nuevo_ancho, nueva_altura))

    wk = Workbook()
    ws = wk.active

    #Obtener Datos de la base de datos
    datos = Incoming.objects.get(sn_batch_pk = sn_batch_pk)


   
    detallesform = Detalle_Incoming.objects.get(incoming_fk=sn_batch_pk)
    
    
# Crea una instancia de la imagen

    image.save('temp_image.png')

# Carga la imagen desde el archivo temporal en OpenPyXL
    excel_image = ExcelImage('temp_image.png')

# Inserta la imagen en la hoja de trabajo
    ws.add_image(excel_image, 'A1')

    ancho_columna = 4.7

    for columna in range(1, 26):
        # Establece el ancho de la columna
        ws.column_dimensions[ws.cell(row=1, column=columna).column_letter].width = ancho_columna


    filas_altura_6 = [3 ,6 ,12 , 16 , 18 , 20 , 22 , 24, 28 , 68]  # Cambia estos valores a las filas que desees

    filas_altura_10_20 = [72]

    # Define la altura predeterminada para las demás filas (12 en este caso)
    altura_predeterminada = 12.60

    # Recorre todas las filas de la hoja
    for fila in ws.iter_rows(min_row=1, max_row=73):
        numero_fila = fila[0].row  # Obtiene el número de fila de la primera celda de la fila
        if numero_fila in filas_altura_6:
            ws.row_dimensions[numero_fila].height = 6
        elif numero_fila in filas_altura_10_20:
            ws.row_dimensions[numero_fila].height = 10.20
        else:
            ws.row_dimensions[numero_fila].height = altura_predeterminada





    border = Border(
        left=Side(border_style="thin"),   # Bordes izquierdos
        right=Side(border_style="thin"),  # Bordes derechos
        top=Side(border_style="thin"),    # Bordes superiores
        bottom=Side(border_style="thin")  # Bordes inferiores
    )


    coordenadas_excepciones = ['R1:T2']

    alineacion = Alignment(horizontal='center', vertical='center')

    rango_celdas = 'A1:Y72'


    ws.merge_cells('R1:T2')
    ws['R1']  = 'RCV N°'
    ws['U1'] = detallesform.rcv_n


    ws.merge_cells('U1:Y2')

    ws.merge_cells('H5:Q5')
    ws['H5']  = 'INSPECCION DE RECEPCION'

    ws.merge_cells('A8:G8')
    ws['A8']  = 'Descripcion'
    ws.merge_cells('A9:G11')#Campo Relleno de BBDD 
    ws['A9'] = datos.descripcion


    ws.merge_cells('H8:O8')
    ws['H8']  = 'N° de Parte/PN'
    ws.merge_cells('H9:O11')#Campo Relleno de BBDD
    ws['H9'] = datos.part_number 

    ws.merge_cells('P8:S8')
    ws['P8']  = 'Modelo'
    ws.merge_cells('P9:S11')#Campo Relleno de BBDD 
    ws['P9'] = detallesform.modelo


    ws.merge_cells('T8:Y8')
    ws['T8']  = 'N° de Serie/SN'
    ws.merge_cells('T9:Y11')#Campo Relleno de BBDD
    
    #PASA EL VALOR SN_BATCH A LA CELDA DEPENDIENDO SI ES SERIAL NUMBER O BATCH NUMBER
    if datos.categoria_fk.categoria_pk == 1:
        ws['T9'] = datos.sn_batch_pk
    elif datos.categoria_fk.categoria_pk == 2:
        ws['E60'] = datos.sn_batch_pk
    elif datos.categoria_fk.categoria_pk == 3:
        ws['T9'] = datos.sn_batch_pk
        ws['E60'] = datos.batch_pk


    
    

    #FALTA CAMBIAR ESPACIO FILA 12

    ws.merge_cells('A13:C13')
    ws['A13']  = 'Overhaul'

    ws.merge_cells('A15:C15')
    ws['A15']  = 'Reparado'

    ws.merge_cells('E13:G13')
    ws['E13']  = 'Chequeado'

    ws.merge_cells('E15:G15')
    ws['E15']  = 'Testeado'

    ws.merge_cells('I13:K13')
    ws['I13']  = 'Otro'

    ws.merge_cells('I15:K15')
    ws['I15']  = 'Nuevo'

    ws.merge_cells('M15:P15')#merge para quitar borde del centro

    ws.merge_cells('M13:O13')
    ws['M13']  = 'Calibración'

    if detallesform.estado_repuesto_fk is None: 
        ws['L13'] = ""
    elif detallesform.estado_repuesto_fk.id == 1: #OVERHAUL
        ws['D13'] = "X"
    elif detallesform.estado_repuesto_fk.id == 2:#REPARADO
        ws['D15'] = "X"
    elif detallesform.estado_repuesto_fk.id == 3:#CHEQUEADO
        ws['H13'] = "X"
    elif detallesform.estado_repuesto_fk.id == 4:#TESTEADO
        ws['H15'] = "X"
    elif detallesform.estado_repuesto_fk.id == 5:#NUEVO
        ws['L15'] = "X"
    elif detallesform.estado_repuesto_fk.id == 6:#CALIBRACION
        ws['P13'] = "X"
    elif detallesform.estado_repuesto_fk.id == 7:#OTRO
        ws['L13'] = "X"
    
    



    ws.merge_cells('Q13:T14')#Campo Relleno de BBDD 
    ws.merge_cells('Q15:T15')
    ws['Q15']  = 'Fecha Recepción'
    ws['Q13'] = datos.f_incoming
    


    ws.merge_cells('U13:W15')
    ws['U13']  = 'Cantidad'
    ws.merge_cells('X13:Y15')#Campo Relleno de BBDD
    ws['X13'] = datos.qty 
    

    ws.merge_cells('A17:G17')
    ws['A17']  = 'Proveedor del Producto'
    ws.merge_cells('H17:P17') #Campo Relleno de BBDD 
    ws['H17'] = detallesform.Proveedor

    ws.merge_cells('Q17:S17')
    ws['Q17']  = 'OC / PO N°'
    ws.merge_cells('T17:Y17') #Campo Relleno de BBDD 
    ws['T17'] = datos.po 


    ws.merge_cells('A19:G19')
    ws['A19']  = 'Taller/Cia. Reparadora'
    ws.merge_cells('H19:P19') #Campo Relleno de BBDD 
    ws['H19'] = detallesform.taller_reparadora

    ws.merge_cells('Q19:S19')
    ws['Q19']  = 'RO N°'
    ws.merge_cells('T19:Y19') #Campo Relleno de BBDD 
    ws['T19'] = detallesform.ro_n

    ws.merge_cells('A21:G21')
    ws['A21']  = 'Trabajo Solicitado'
    ws.merge_cells('H21:P21') #Campo Relleno de BBDD 
    ws['H21'] = detallesform.trabajo_solicitado

    ws.merge_cells('Q21:S21')
    ws['Q21']  = 'WO N°'
    ws.merge_cells('T21:Y21') #Campo Relleno de BBDD 
    ws['T21'] = detallesform.wo_n

    ws.merge_cells('A23:G23')
    ws['A23']  = 'Propiedad de'
    ws.merge_cells('H23:Y23') #Campo Relleno de BBDD 
    ws['H23'] = detallesform.propiedad

    ws.merge_cells('A25:G25')
    ws['A25']  = 'Check Periodica (Si Aplica)'
    ws.merge_cells('H25:N25') #Campo Relleno de BBDD 
    ws['H25'] = detallesform.check_periodica


    ws.merge_cells('O25:R25')
    ws['O25']  = 'Shel Life Due'
    ws.merge_cells('S25:Y25') #Campo Relleno de BBDD 
    ws['S25'] = datos.f_vencimiento 

    #ENCABEZADO DE LA TABLA
    ws['A27'] = 'Item'
    ws.merge_cells('B27:S27')
    ws['B27']  = 'Materia'
    ws.merge_cells('T27:V27') 
    ws['T27']  = 'SI'
    ws.merge_cells('W27:Y27') 
    ws['W27']  = 'NO'


    if detallesform.item1 == '0':
        ws['W29'] = "X"
    elif detallesform.item1 == '1':
        ws['T29'] = "X"


    if detallesform.item2 == '0': #NO
        ws['W30'] = "X"
    elif detallesform.item2 == '1':#SI
        ws['T30'] = "X"

    if detallesform.item3 == '0': #NO
        ws['W31'] = "X"
    elif detallesform.item3 == '1':#SI
        ws['T31'] = "X"

    if detallesform.item4 == '0': #NO
        ws['W32'] = "X"
    elif detallesform.item4 == '1':#SI
        ws['T32'] = "X"

    if detallesform.item5 == '0': #NO
        ws['W33'] = "X"
    elif detallesform.item5 == '1':#SI
        ws['T33'] = "X"

    if detallesform.item6 == '0': #NO
        ws['W34'] = "X"
    elif detallesform.item6 == '1':#SI
        ws['T34'] = "X"
    
    if detallesform.item7 == '0': #NO
        ws['W35'] = "X"
    elif detallesform.item7 == '1':#SI
        ws['T35'] = "X"

    if detallesform.item8 == '0': #NO
        ws['W36'] = "X"
    elif detallesform.item8 == '1':#SI
        ws['T36'] = "X"

    if detallesform.item9 == '0': #NO
        ws['W37'] = "X"
    elif detallesform.item9 == '1':#SI
        ws['T37'] = "X"
    
    if detallesform.item10 == '0': #NO
        ws['W38'] = "X"
    elif detallesform.item10 == '1':#SI
        ws['T38'] = "X"

    if detallesform.item11 == '0': #NO
        ws['W39'] = "X"
    elif detallesform.item11 == '1':#SI
        ws['T39'] = "X"

    if detallesform.item12 == '0': #NO
        ws['W40'] = "X"
    elif detallesform.item12 == '1':#SI
        ws['T40'] = "X"

    if detallesform.item13 == '0': #NO
        ws['W41'] = "X"
    elif detallesform.item13 == '1':#SI
        ws['T41'] = "X"
        ws['P41'] = detallesform.n_item13

    if detallesform.item14 == '0': #NO
        ws['W42'] = "X"
    elif detallesform.item14 == '1':#SI
        ws['T42'] = "X"

    if detallesform.item15 == '0': #NO
        ws['W43'] = "X"
    elif detallesform.item15 == '1':#SI
        ws['T43'] = "X"
        ws['M43'] = detallesform.n_item15
    
    if detallesform.item16 == '0': #NO
        ws['W44'] = "X"
    elif detallesform.item16 == '1':#SI
        ws['T44'] = "X"
        ws['M44'] = detallesform.n_item16

    if detallesform.item17 == '0': #NO
        ws['W45'] = "X"
    elif detallesform.item17 == '1':#SI
        ws['T45'] = "X"
        ws['M45'] = detallesform.n_item17

    if detallesform.item18 == '0': #NO
        ws['W46'] = "X"
    elif detallesform.item18 == '1':#SI
        ws['T46'] = "X"
        if detallesform.item_18tsn == "TSN":
            ws['I46'] = detallesform.n_item18tsn
        elif detallesform.item_18tsn == "TSO":
            ws['L46'] = detallesform.n_item18tsn
        elif detallesform.item_18tsn == "CSN":
            ws['O46'] = detallesform.n_item18tsn
        elif detallesform.item_18tsn == "CS8":
            ws['R46'] = detallesform.n_item18tsn


    if detallesform.item19 == '0': #NO
        ws['W47'] = "X"
    elif detallesform.item19 == '1':#SI
        ws['T47'] = "X"
    
    
    if detallesform.item20 == '0': #NO
        ws['W48'] = "X"
    elif detallesform.item20 == '1':#SI
        ws['T48'] = "X"

    if detallesform.item21 == '0': #NO
        ws['W49'] = "X"
    elif detallesform.item21 == '1':#SI
        ws['T49'] = "X"

    if detallesform.item22 == '0': #NO
        ws['W50'] = "X"
    elif detallesform.item22 == '1':#SI
        ws['T50'] = "X"
        ws['M50'] = detallesform.n_item22




    #CUERPO DE LA TABLA
    ws['A29'] = '1'
    ws.merge_cells('B29:S29')
    ws['B29'] = 'Producto conforme a lo indicado en la lista de Embarque (Packing List)'
    ws.merge_cells('T29:V29')
    ws.merge_cells('W29:Y29')

    ws['A30'] = '2'
    ws.merge_cells('B30:S30')
    ws['B30'] = 'Factura del Proveedor conforme a la orden de compra o solicitud de trabajo (Invoice)'
    ws.merge_cells('T30:V30')
    ws.merge_cells('W30:Y30')

    ws['A31'] = '3'
    ws.merge_cells('B31:S31')
    ws['B31'] = 'Cartilla/Orden de Trabajo, Cartilla de prueba, Si corresponde, del taller que repara'
    ws.merge_cells('T31:V31')
    ws.merge_cells('W31:Y31')

    ws['A32'] = '4'
    ws.merge_cells('B32:S32')
    ws['B32'] = 'Producto sin daños visibles (Inspeccion Visual)'
    ws.merge_cells('T32:V32')
    ws.merge_cells('W32:Y32')


    ws['A33'] = '5'
    ws.merge_cells('B33:S33')
    ws['B33'] = 'Producto protegido en embalaje apropiado'
    ws.merge_cells('T33:V33')
    ws.merge_cells('W33:Y33')

    ws['A34'] = '6'
    ws.merge_cells('B34:S34')
    ws['B34'] = 'Placa de identificacion del componente'
    ws.merge_cells('T34:V34')
    ws.merge_cells('W34:Y34')

    ws['A35'] = '7'
    ws.merge_cells('B35:S35')
    ws['B35'] = 'Documentacion Técnica completa requerida por reglamentacion (Trazabilidad)'
    ws.merge_cells('T35:V35')
    ws.merge_cells('W35:Y35')

    ws['A36'] = '8'
    ws.merge_cells('B36:S36')
    ws['B36'] = 'Formulario FAA 8130-3'
    ws.merge_cells('T36:V36')
    ws.merge_cells('W36:Y36')

    ws['A37'] = '9'
    ws.merge_cells('B37:S37')
    ws['B37'] = 'Formulario EASA Form One o JAA Form One o CAA Form One'
    ws.merge_cells('T37:V37')
    ws.merge_cells('W37:Y37')

    ws['A38'] = '10'
    ws.merge_cells('B38:S38')
    ws['B38'] = 'Formulario DGAC Chile 8130-3'
    ws.merge_cells('T38:V38')
    ws.merge_cells('W38:Y38')

    ws['A39'] = '11'
    ws.merge_cells('B39:S39')
    ws['B39'] = 'Formulario ANAC Argentina 8130-3																	'
    ws.merge_cells('T39:V39')
    ws.merge_cells('W39:Y39')

    ws['A40'] = '12'
    ws.merge_cells('B40:S40')
    ws['B40'] = 'Placa o Etiqueta para Herramientas o equipos con calibracion'
    ws.merge_cells('T40:V40')
    ws.merge_cells('W40:Y40')

    ws['A41'] = '13'
    ws.merge_cells('B41:N41')
    ws['B41'] = 'Certificado de Calibracion en laboratorio reconocido por el estado local'
    ws['O41'] = 'N°'
    ws.merge_cells('P41:S41')
    ws.merge_cells('T41:V41')
    ws.merge_cells('W41:Y41')


    ws['A42'] = '14'
    ws.merge_cells('B42:S42')
    ws['B42'] = 'Materiales con vida limite (Verificacion de Shelf life data y MSDS)'
    ws.merge_cells('T42:V42')
    ws.merge_cells('W42:Y42')

    ws['A43'] = '15'
    ws.merge_cells('B43:K43')
    ws['B43'] = 'Certificado de flamabilidad, si corresponde'
    ws.merge_cells('M43:S43')#CAMPO LLENADO POR BBDD
    ws['L43'] = 'N°'
    ws.merge_cells('T43:V43')
    ws.merge_cells('W43:Y43')

    ws['A44'] = '16'
    ws.merge_cells('B44:K44')
    ws['B44'] = 'Certificado de conformidad  y/o Analisis'
    ws.merge_cells('M44:S44')#CAMPO LLENADO POR BBDD
    ws['l44'] = 'N°'
    ws.merge_cells('T44:V44')
    ws.merge_cells('W44:Y44')

    ws['A45'] = '17'
    ws.merge_cells('B45:K45')
    ws['B45'] = 'Numero de lote de fabricacion, si corresponde'
    ws.merge_cells('M45:S45')#CAMPO LLENADO POR BBDD
    ws['L45'] = 'N°'
    ws.merge_cells('T45:V45')
    ws.merge_cells('W45:Y45')

    ws['A46'] = '18'
    ws.merge_cells('B46:G46')
    ws['B46'] = 'TSO / TSN (Si Aplica)'

    ws.merge_cells('I46:J46')#CAMPO LLENADO POR BBDD
    ws['H46'] = 'TSN'

    ws.merge_cells('L46:M46')
    ws['K46'] = 'TSO'

    ws.merge_cells('O46:P46')
    ws['N46'] = 'CSN'

    ws.merge_cells('R46:S46')
    ws['Q46'] = 'CSO'

    ws.merge_cells('T46:V46')
    ws.merge_cells('W46:Y46')

    ws['A47'] = '19'
    ws.merge_cells('B47:S47')
    ws['B47'] = 'Material Safety Data Sheet'
    ws.merge_cells('T47:V47')
    ws.merge_cells('W47:Y47')

    ws['A48'] = '20'
    ws.merge_cells('B48:S48')
    ws['B48'] = 'Material con restriccion bajo el programa ESD'
    ws.merge_cells('T48:V48')
    ws.merge_cells('W48:Y48')

    ws['A49'] = '21'
    ws.merge_cells('B49:S49')
    ws['B49'] = 'Material con restriccion de almacenamiento (Hielo Seco)'
    ws.merge_cells('T49:V49')
    ws.merge_cells('W49:Y49')

    ws['A50'] = '22'
    ws.merge_cells('B50:K50')
    ws['B50'] = 'Cartilla Mantencion CMA Autorizado'

    ws.merge_cells('M50:S50')
    ws['L50'] = 'N°'

    ws.merge_cells('T50:V50')
    ws.merge_cells('W50:Y50')

    ws.merge_cells('A52:Y52')
    ws['A52'] = 'Observaciones'
    ws.merge_cells('A53:Y56')
    ws['A53'] = datos.observaciones

    ws.merge_cells('A58:Y58')
    ws['A58'] = 'Observaciones del Producto Recepcionado'

    ws.merge_cells('A59:D59')
    ws['A59'] = 'OC, PO OR RO N°'
    ws.merge_cells('E59:Y59')
    ws['E59'] = datos.po


    ws.merge_cells('A60:D60')
    ws['A60'] = 'BATCH'
    ws.merge_cells('E60:Y60')

    ws.merge_cells('A61:D61')
    ws['A61'] = 'STDF'
    ws.merge_cells('E61:Y61')
    ws['E61'] = datos.stdf_fk.stdf_pk 

    ws.merge_cells('A62:D62')
    ws['A62'] = 'AWB'
    ws.merge_cells('E62:Y62')
    ws['E62'] = datos.stdf_fk.awb 

    ws.merge_cells('A63:D64')
    ws['A63'] = 'UBICACION'
    ws.merge_cells('E63:Y64')
    ws['E63'] = datos.ubicacion_fk.name_ubicacion 
    

    ws.merge_cells('A65:D65')
    ws['A65'] = 'ACEPTADO'

    

    ws.merge_cells('E65:F65')
    ws['E65'] = 'SI'
    ws.merge_cells('G65:H65')

    ws.merge_cells('I65:J65')
    ws['I65'] = 'NO'
    ws.merge_cells('K65:L65')

    ws.merge_cells('M65:Y65')

    if detallesform.aceptado == 'SI':
        ws['G65'] = 'X'
    elif detallesform.aceptado == 'NO':
        ws['K65'] = 'X'


    ws.merge_cells('A67:H67')
    ws['A67'] = 'Nombre'

    ws.merge_cells('I67:P67')
    ws['I67'] = 'N° Licencia'
    ws['I69'] = detallesform.licencia.name_licencia
    

    ws.merge_cells('Q67:Y67')
    ws['Q67'] = 'Firma'

    ws.merge_cells('A69:H71')
    nombre_completo = f"{datos.usuario.first_name} {datos.usuario.last_name}"
    ws['A69'] = nombre_completo
    
    ws.merge_cells('I69:P71')


    ws.merge_cells('Q69:Y71') #FIRMA DEL USUARIO 

    ws.merge_cells('A72:J72')
    ws['A72'] = 'FORM CMA-005 REV.1'

    #ws.merge_cells('')
    #ws['']  = ''
    #ws.merge_cells(':') #Campo Relleno de BBDD 

    coordenadas_excepciones = ['A13:C13']
    coordenadas_combinadas = ws.merged_cells.ranges
    for rango in coordenadas_combinadas:
        for row in ws.iter_rows(min_row=rango.min_row, max_row=rango.max_row,
                            min_col=rango.min_col, max_col=rango.max_col):
            for cell in row:
                if any(coord in cell.coordinate for coord in coordenadas_excepciones):
                    continue  # Salta las celdas en coordenadas_excepciones
                cell.border = border


    coordenadas_a_quitar_borde = [('A72','J72'),('R1', 'T2'),('H5', 'Q5'),('A13', 'C13'), ('E13', 'G13'), ('I13', 'K13'), ('M13', 'P13'), ('A15', 'C15'), ('E15', 'G15'), ('I15', 'K15')]

    # Iterar sobre las coordenadas y quitar el borde
    for coord in coordenadas_a_quitar_borde:
        min_coord, max_coord = coord

        # Convertir las coordenadas en números enteros
        min_row, min_col = int(min_coord[1:]), ord(min_coord[0]) - ord('A') + 1
        max_row, max_col = int(max_coord[1:]), ord(max_coord[0]) - ord('A') + 1

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = None


    for fila in ws.iter_rows(min_row=1, max_row=72, min_col=1, max_col=25):
        for cell in fila:
            cell.alignment = alineacion



    # Configura la escala de la hoja para ajustarla a una página
    ws.page_setup.fitToPage = True

    # Configura los márgenes de la página para reducirlos si es necesario
    ws.page_margins.left = 0.25  # Márgen izquierdo
    ws.page_margins.right = 0.30  # Márgen derecho
    ws.page_margins.top = 0.25  # Márgen superior
    ws.page_margins.bottom = 0.25  # Márgen inferior


    for row in ws.iter_rows(min_row=13, max_row=15, min_col=1, max_col=16):  # Columna A a P
        for cell in row:
            if cell.row == 13 or cell.row == 15 or cell.column == 'A' or cell.column == 'P':
                cell.border = border

    # Eliminar los bordes en las celdas del interior del rango
    for row in ws.iter_rows(min_row=14, max_row=14, min_col=2, max_col=15):  # Excluye las celdas del perímetro
        for cell in row:
            cell.border = None


    coordenadas_a_bordear = ['A27','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38','A39','A40','A41','A42',
                            'A43','A44','A45','A46','A47','A48','A49','A50']
    # Aplicar el borde a las celdas con un ciclo for
    for coord in coordenadas_a_bordear:
        celda = ws[coord]
        celda.border = border

    # Define el formato de fuente para las demás filas
    borde_sin_superior = Border(top=Side(style=None))

    # Celda combinada 'M15:P15'
    celda_combinada = ws['M15:P15']

    # Aplica el estilo de borde sin borde superior a la celda combinada
    for fila in celda_combinada:
        for celda in fila:
            celda.border = borde_sin_superior


    borde_izquierdo = Border(left=Side(style='thin'),
                            top=Side(style=None),
                            right=Side(style=None),
                            bottom=Side(style=None))

    # Aplica el estilo de borde izquierdo a la celda A14
    celda_a14 = ws['A14']
    celda_a14.border = borde_izquierdo


    borde_abajo = Border(left=Side(style=None),
                        top=Side(style='thin'),
                        right=Side(style=None),
                        bottom=Side(style=None))

    # Celdas combinadas
    celdas_combinadas1 = ws['I13:K13']
    celdas_combinadas2 = ws['A13:C13']
    celdas_combinadas3 = ws['E13:G13']
    celdas_combinadas4 = ws['M13:P13']

    # Lista de celdas combinadas
    celdas_combinadas = [celdas_combinadas2, celdas_combinadas3, celdas_combinadas4, celdas_combinadas1]

    # Aplica el estilo de borde inferior a todas las celdas combinadas
    for rango_celdas in celdas_combinadas:
        for fila in rango_celdas:
            for celda in fila:
                celda.border = borde_abajo




    borde_arriba = Border(left=Side(style=None),
                        top=Side(style=None),
                        right=Side(style=None),
                        bottom=Side(style='thin'))

    # Celdas combinadas
    celdas_combi1 = ws['A15:C15']
    celdas_combi2 = ws['E15:G15']
    celdas_combi3 = ws['I15:K15']
    celdas_combi4 = ws['M15:P15']


    # Lista de celdas combinadas
    celdas_combi = [celdas_combi1, celdas_combi2, celdas_combi3,celdas_combi4]

    # Aplica el estilo de borde inferior a todas las celdas combinadas
    for i in celdas_combi:
        for a in i:
            for celda in a:
                celda.border = borde_arriba

    borde_izquierdo_abajo = Border(left=Side(style='thin'),
                            top=Side(style='thin'),
                            right=Side(style=None),
                            bottom=Side(style='thin'))

    # Aplica el estilo de borde izquierdo a la celda A14
    celda_P13 = ws['P13']
    celda_P13.border = borde_izquierdo_abajo

    alineacion_izquierda = Alignment(horizontal='left')

    # Celdas combinadas
    celdas_combinadas1 = ws['E59:Y59']
    celdas_combinadas2 = ws['E60:Y60']
    celdas_combinadas3 = ws['E61:Y61']
    celdas_combinadas4 = ws['E62:Y62']
    celdas_combinadas5 = ws['E63:Y64']

    # Lista de celdas combinadas que se alinearán a la izquierda
  

    inicio_fila = 1
    final_fila = 72
    columnas = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'  # Columnas de la A a la Z

    # Crear un estilo con alineación vertical al centro
    
    centrar_vertical = Alignment(wrapText=True, vertical='center',horizontal='center')

    # Iterar a través de las celdas y ajustar el texto y la alineación vertical
    for fila in range(inicio_fila, final_fila + 1):
        for columna in columnas:
            celda = ws[columna + str(fila)]
            celda.alignment = centrar_vertical
            

    celdas_combinadas1 = ws['E59:Y59']
    celdas_combinadas2 = ws['E60:Y60']
    celdas_combinadas3 = ws['E61:Y61']
    celdas_combinadas4 = ws['E62:Y62']
    celdas_combinadas5 = ws['E63:Y64']

    celdas_combinadas_izquierda = [celdas_combinadas1, celdas_combinadas2, celdas_combinadas3, celdas_combinadas4, celdas_combinadas5]

    # Aplica la alineación a la izquierda a todas las celdas combinadas
    for rango_celdas in celdas_combinadas_izquierda:
        for fila in rango_celdas:
            for celda in fila:
                celda.alignment = alineacion_izquierda
    rango_inicio = 'B29'
    rango_fin = 'S50'

    # Crear un objeto Alignment para alinear a la izquierda
    alineacion_izquierda = Alignment(horizontal='left')

    # Aplicar la alineación a todas las celdas en el rango
    for row in ws.iter_rows(min_row=29, max_row=50, min_col=2, max_col=19):  # Columna B a S
        for cell in row:
            cell.alignment = alineacion_izquierda
    
    formato_fuente = Font(name='Tahoma', size=9 , bold=True)

    # Define el formato de fuente para la fila 72
    formato_fuente_fila_72 = Font(name='Tahoma', size=7, bold=True)
    alineacion_izquierda = Alignment(horizontal='left')

    # Itera a través de las filas
    for fila in ws.iter_rows(min_row=1, max_row=74, min_col=1, max_col=25):
        for celda in fila:
            # Aplica el formato de fuente para todas las celdas
            celda.font = formato_fuente

    # Aplica negrita a todas las celdas, incluyendo la fila 72
    for celda in ws[72]:
        celda.font = formato_fuente_fila_72
        celda.alignment = alineacion_izquierda
    

    wk.save(response)
    return response