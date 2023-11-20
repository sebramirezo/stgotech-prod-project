from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , Border, Side, PatternFill
from .forms import *
from django.http import HttpResponse
from .models import *
from openpyxl.utils import get_column_letter

def inventario_comat(request):
    comat_data = Comat.objects.all()

    wk = Workbook()
    ws = wk.active


    encabezado = ['RESPONSABLE', 'CORRELATIVO', 'AWB',' HAWB', 'BODEGA', 'ORIGEN', 'PCS', 'PESO', 'STDF', 'FECHA DE MANIIFESTO', 'FECHA DE CONTROL','FECHA DE RECEPCION', 'OBSERVACIONES'] 
    ws.append(encabezado)

    # Añade datos a las columnas siguientes
    for comat in comat_data:
        data_column = [
            comat.usuario.username,
            comat.corr_interno, 
            comat.awb, 
            comat.hawb, 
            comat.bodega_fk.name_bodega, 
            comat.origen_fk.name_origen, 
            comat.pcs, comat.peso, 
            comat.stdf_pk, 
            comat.f_manifiesto.replace(tzinfo=None) if comat.f_manifiesto else None , 
            comat.f_control.replace(tzinfo=None) if comat.f_control else None, 
            comat.f_recepcion.replace(tzinfo=None) if comat.f_recepcion else None, 
            comat.observaciones]

        # Transpone la columna a una fila
        ws.append(data_column)

    # Crea una respuesta HTTP con el contenido del libro de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ControlAWBSTGO.xlsx'

    for column in range(1, 14):  # Columnas de la A a la M
        col_letter = get_column_letter(column)
        ws.column_dimensions[col_letter].width = 20

    # Agrega bordes a las celdas desde A hasta M y hasta la última fila
    max_row = ws.max_row
    max_col_letter = get_column_letter(13)  # Columna M

    for row in range(1, max_row + 1):
        for col in range(1, 14):  # Columnas de la A a la M
            col_letter = get_column_letter(col)
            
            # Define los bordes para la celda actual
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplica los bordes a la celda
            ws[col_letter + str(row)].border = border_style

            # Centra el contenido en la celda
            ws[col_letter + str(row)].alignment = Alignment(horizontal='center', vertical='center')

            # Cambia el color de fondo, el color de la letra y pone en negrita para la primera fila
            if row == 1:
                ws[col_letter + str(row)].fill = PatternFill(start_color='073763', end_color='073763', fill_type='solid')  # Azul
                ws[col_letter + str(row)].font = Font(color="FFFFFF", bold=True)  # Blanco y negrita


    # Guarda ellibro de Excel en la respuesta HTTP
    wk.save(response)

    return response


def inventario_incoming(request):
    # Obtén todos los objetos Incoming del modelo
    incoming_data = Incoming.objects.all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active

    # Añade encabezados a la primera columna de la hoja
    header_row = [
        'Owner',
        'Ubicacion',
        'Part Number',
        'Descripcion',
        'Batch Number',
        'Serial Number',
        'Qty',
        'Saldo',
        'UOM',
        'Fecha Vencimiento',
        'Clasificacion',
        'STDF',
        'Condicion',
        'Fecha Incoming',
    ]

    ws.append(header_row)

    # Ajusta el ancho de las celdas a 20 pixeles
    for column in range(1, len(header_row) + 1):
        col_letter = get_column_letter(column)
        ws.column_dimensions[col_letter].width = 20

    # Agrega bordes a las celdas y configura el estilo de la primera fila
    for row in range(1, 2):  # Solo la primera fila
        for col in range(1, len(header_row) + 1):
            col_letter = get_column_letter(col)

            # Define los bordes para la celda actual
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplica los bordes a la celda
            ws[col_letter + str(row)].border = border_style

            # Centra el contenido en la celda
            ws[col_letter + str(row)].alignment = Alignment(horizontal='center', vertical='center')

            # Cambia el color de fondo, el color de la letra y pone en negrita para la primera fila
            ws[col_letter + str(row)].fill = PatternFill(start_color='ffc000', end_color='ffc000', fill_type='solid')  # Azul
            ws[col_letter + str(row)].font = Font(color="FFFFFF", bold=True)  # Blanco y negrita

    # Agrega datos a las columnas siguientes
    for incoming in incoming_data:
        data_column = []
        # Asegúrate de que los objetos de fecha y hora no tengan información de zona horaria
        if incoming.categoria_fk.categoria_pk == 1:
            data_column = [
                incoming.owner_fk.name_owner,
                incoming.ubicacion_fk.name_ubicacion, 
                incoming.part_number, 
                incoming.descripcion, 
                None,  
                incoming.sn_batch_pk,  # Serial Number
                incoming.qty, 
                incoming.saldo, 
                incoming.uom_fk.name_uom, 
                incoming.f_vencimiento, 
                incoming.clasificacion_fk.name_clasificacion, 
                incoming.stdf_fk.stdf_pk,
                incoming.condicion_fk.name_condicion,
                incoming.f_incoming,
            ]
        elif incoming.categoria_fk.categoria_pk == 2:
            data_column = [
                incoming.owner_fk.name_owner,
                incoming.ubicacion_fk.name_ubicacion, 
                incoming.part_number, 
                incoming.descripcion, 
                incoming.sn_batch_pk, 
                None,  
                incoming.qty, 
                incoming.saldo, 
                incoming.uom_fk.name_uom, 
                incoming.f_vencimiento, 
                incoming.clasificacion_fk.name_clasificacion, 
                incoming.stdf_fk.stdf_pk,
                incoming.condicion_fk.name_condicion,
                incoming.f_incoming,
            ]
        elif incoming.categoria_fk.categoria_pk == 3:
            data_column = [
                incoming.owner_fk.name_owner,
                incoming.ubicacion_fk.name_ubicacion, 
                incoming.part_number, 
                incoming.descripcion, 
                incoming.batch_pk, 
                incoming.sn_batch_pk,  
                incoming.qty, 
                incoming.saldo, 
                incoming.uom_fk.name_uom, 
                incoming.f_vencimiento, 
                incoming.clasificacion_fk.name_clasificacion, 
                incoming.stdf_fk.stdf_pk,
                incoming.condicion_fk.name_condicion,
                incoming.f_incoming,
            ]

        # Transpone la columna a una fila
        ws.append(data_column)

    # Crea una respuesta HTTP con el contenido del libro de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=incoming_data.xlsx'

    # Guarda el libro de Excel en la respuesta HTTP
    wb.save(response)

    return response

